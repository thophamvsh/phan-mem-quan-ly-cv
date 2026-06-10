import os
import io
import json
from django.http import HttpResponse, JsonResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
from datetime import datetime, timedelta
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from core.factory_scope import filter_queryset_by_factory, get_user_factory_name, get_user_factory_code, has_profile_permission, has_all_factory_access
from .models import ThongSoVanHanh, ThietBi


def _looks_like_index_or_time_column(series, expected_rows):
    """
    Detect the leading helper column used by the newer template.

    Legacy Song Hinh templates put the first parameter directly in column A.
    Newer templates add an STT/time column in A and start parameters in B.
    """
    values = [value for value in series.head(expected_rows).tolist() if pd.notna(value)]
    if not values:
        return False

    numeric_values = []
    for value in values:
        if isinstance(value, str) and ":" in value:
            try:
                datetime.strptime(value.strip(), "%H:%M")
                continue
            except ValueError:
                return False

        try:
            numeric_values.append(int(float(value)))
        except (TypeError, ValueError):
            return False

    if not numeric_values:
        return True

    expected_sequence = list(range(1, len(numeric_values) + 1))
    zero_based_sequence = list(range(len(numeric_values)))
    return numeric_values in (expected_sequence, zero_based_sequence)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def excel_template(request):
    """
    Tạo template Excel cho thông số vận hành điện theo cấu hình động của nhà máy
    """
    try:
        if not has_profile_permission(request.user, "can_export_excel"):
            return JsonResponse({'error': 'Tài khoản của bạn chưa được cấp quyền xuất dữ liệu hoặc tải template Excel. Vui lòng liên hệ quản trị viên.'}, status=403)
            
        factory_code = request.query_params.get("factory_code")
        if not factory_code or not has_all_factory_access(request.user):
            factory_code = get_user_factory_code(request.user) or 'SH'
        from quanlyvanhanh.configs.operation_configs import get_dien_factory_config
        config = get_dien_factory_config(factory_code)
        
        # Tạo workbook mới
        wb = Workbook()
        ws = wb.active
        ws.title = "ThongSoVanHanhDien"
        
        # Tính toán headers động
        device_headers = []
        parameter_headers = []
        unit_headers = []
        
        for grp in config['layout']:
            num_cols = len(grp['columns'])
            # Tên nhóm thiết bị ở ô đầu tiên của nhóm, các ô sau để trống để gộp ô
            device_headers.append(grp['group'])
            device_headers.extend([""] * (num_cols - 1))
            
            for col in grp['columns']:
                parameter_headers.append(col['ten'])
                unit_headers.append(col['don_vi'])
                
        # Tổng số cột thông số
        total_cols = len(parameter_headers)
        
        # Ghi các hàng header
        if factory_code == 'VS':
            # Row 1: Ngày, date, Title
            ws.cell(row=1, column=3, value="Ngày").font = Font(name="Times New Roman", size=12, bold=True, color="FF0000")
            ws.cell(row=1, column=3).alignment = Alignment(horizontal='center', vertical='center')
            
            current_date_str = timezone.localtime(timezone.now()).strftime("%d-%m-%Y")
            ws.cell(row=1, column=5, value=current_date_str).font = Font(name="Times New Roman", size=12, bold=True, color="FF0000")
            ws.cell(row=1, column=5).alignment = Alignment(horizontal='center', vertical='center')
            
            title_cell = ws.cell(row=1, column=12, value="BẢNG NHẬP THÔNG SỐ VẬN HÀNH ĐIỆN")
            title_cell.font = Font(name="Times New Roman", size=14, bold=True, color="0000FF")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells("L1:AK1")
            
            ws.merge_cells("A2:A3")
            loc_ked = ws.cell(row=2, column=1, value="LOC\nKED")
            loc_ked.font = Font(name="Times New Roman", size=12, bold=True)
            loc_ked.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            fills_map = {
                'MÁY PHÁT H1': 'E2EFDA',
                'MÁY PHÁT H2': 'E2EFDA',
                'TRẠM PHÂN PHỐI 110 kV': 'F2DCDB',
                'MBA T1, T2': 'FFFFFF',  # White for group header
                'Áp suất khí máy cắt 110kV': 'FCE4D6',
                'MBA tự dùng TD91': 'E2EFDA',
                'MBA tự dùng TD92': 'F8CBAD',
            }
            
            current_col = 2
            for grp in config['layout']:
                num_cols = len(grp['columns'])
                grp_name = grp['group']
                
                # Style row 2 (Group header)
                group_color = fills_map.get(grp_name, 'D9E1F2')
                group_fill = PatternFill(start_color=group_color, end_color=group_color, fill_type='solid')
                
                for col_idx in range(current_col, current_col + num_cols):
                    cell = ws.cell(row=2, column=col_idx)
                    cell.font = Font(name="Times New Roman", size=12, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.fill = group_fill
                    
                if num_cols > 1:
                    start_letter = get_column_letter(current_col)
                    end_letter = get_column_letter(current_col + num_cols - 1)
                    ws.merge_cells(f"{start_letter}2:{end_letter}2")
                    
                # Style row 3 (Parameter header)
                for i, col in enumerate(grp['columns']):
                    col_idx = current_col + i
                    cell = ws.cell(row=3, column=col_idx, value=col['ten'])
                    cell.font = Font(name="Times New Roman", size=12, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    if grp_name == 'MBA T1, T2':
                        if col.get('sub_device') == 'T1':
                            param_color = 'E2EFDA'
                        else:
                            param_color = 'DDEBF7'
                    else:
                        param_color = group_color
                        
                    cell.fill = PatternFill(start_color=param_color, end_color=param_color, fill_type='solid')
                    
                current_col += num_cols
                
            row_index = 4
            
        else: # Sông Hinh
            ws.merge_cells("A1:A3")
            for r in range(1, 4):
                cell = ws.cell(row=r, column=1)
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            stt_cell = ws.cell(row=1, column=1, value="STT")
            stt_cell.font = Font(bold=True)
            stt_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for col_idx, header in enumerate(device_headers, 2):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                
            for col_idx, header in enumerate(parameter_headers, 2):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
                
            for col_idx, header in enumerate(unit_headers, 2):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = PatternFill(start_color='DDA0DD', end_color='DDA0DD', fill_type='solid')
                
            current_col = 2
            for grp in config['layout']:
                num_cols = len(grp['columns'])
                if num_cols > 1:
                    start_letter = get_column_letter(current_col)
                    end_letter = get_column_letter(current_col + num_cols - 1)
                    ws.merge_cells(f"{start_letter}1:{end_letter}1")
                current_col += num_cols
                
            row_index = 4

        # Điền STT và chu kỳ dữ liệu trống
        slots = []
        if factory_code == 'VS':
            for hour in range(24):
                slots.append(f"{hour:02d}:00")
        else:
            for hour in range(24):
                slots.append(f"{hour:02d}:00")
                slots.append(f"{hour:02d}:30")
                
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') if factory_code == 'VS' else None
        for idx, slot in enumerate(slots, 1):
            stt_val_cell = ws.cell(row=row_index, column=1, value=idx)
            if factory_code == 'VS':
                stt_val_cell.font = Font(name="Times New Roman", size=12, bold=True)
            else:
                stt_val_cell.font = Font(bold=True)
            stt_val_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for col in range(2, total_cols + 2):
                cell = ws.cell(row=row_index, column=col, value="")
                if factory_code == 'VS':
                    cell.font = Font(name="Times New Roman", size=12)
                    if yellow_fill:
                        cell.fill = yellow_fill
            row_index += 1
            
        # Điều chỉnh độ rộng cột
        ws.column_dimensions['A'].width = 8
        for col in range(2, total_cols + 2):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 12
            
        # Kẻ viền cho toàn bộ bảng
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for r in range(1, 4 + len(slots)):
            for col in range(1, total_cols + 2):
                if factory_code == 'VS' and r == 1:
                    continue
                ws.cell(row=r, column=col).border = thin_border
                
        # Tạo response với buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ThongSoVanHanhDien_Template_{factory_code}.xlsx"'
        response['Content-Length'] = len(buffer.getvalue())
        return response
        
    except Exception as e:
        import traceback
        error_msg = f"Lỗi tạo template Excel: {str(e)}\n{traceback.format_exc()}"
        return JsonResponse({'error': error_msg}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def excel_import(request):
    """
    Import dữ liệu từ file Excel động theo cấu hình nhà máy
    """
    try:
        if not has_profile_permission(request.user, "can_import_excel"):
            return JsonResponse({'error': 'Tài khoản của bạn chưa được cấp quyền import dữ liệu từ Excel. Vui lòng liên hệ quản trị viên.'}, status=403)
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'Không có file được upload'}, status=400)

        file = request.FILES['file']

        # Lấy ngày và thời điểm từ form
        request_data = getattr(request, "data", request.POST)
        import_date = request_data.get('selected_date')
        time_slots_json = request_data.get('time_slots', '[]')

        if not import_date:
            return JsonResponse({'error': 'Vui lòng chọn ngày import'}, status=400)

        # Parse time slots
        try:
            time_slots = json.loads(time_slots_json)
        except Exception:
            time_slots = []

        # Tự động nhận diện nhà máy từ file Excel
        try:
            temp_df = pd.read_excel(file, nrows=3, header=None)
            first_rows_text = ""
            for r_idx in range(min(3, len(temp_df))):
                row_vals = [str(x) for x in temp_df.iloc[r_idx].tolist() if pd.notna(x)]
                first_rows_text += " " + " ".join(row_vals)
            first_rows_text = first_rows_text.upper()
            num_cols = len(temp_df.columns)
        except Exception:
            first_rows_text = ""
            num_cols = 0

        detected_factory = 'SH'
        if num_cols >= 35:
            if "SÔNG HINH" in first_rows_text or "SONG HINH" in first_rows_text:
                detected_factory = 'SH'
            else:
                detected_factory = 'VS'
        else:
            detected_factory = 'SH'

        # Lấy cấu hình nhà máy và số chu kỳ
        factory_code = request_data.get('factory_code')
        if not factory_code or not has_all_factory_access(request.user):
            factory_code = get_user_factory_code(request.user)
            
        if not factory_code:
            factory_code = detected_factory
        num_cycles = 24 if factory_code == 'VS' else 48

        # Tạo time cycles list trước khi đọc Excel
        if time_slots and factory_code != 'VS':
            time_cycles_list = [slot['time'] for slot in time_slots if slot.get('selected', True)]
        else:
            time_cycles_list = []
            if factory_code == 'VS':
                for hour in range(24):
                    time_cycles_list.append(f"{hour:02d}:00")
            else:
                for hour in range(24):
                    time_cycles_list.append(f"{hour:02d}:00")
                    time_cycles_list.append(f"{hour:02d}:30")

        from quanlyvanhanh.configs.operation_configs import get_dien_factory_config
        config = get_dien_factory_config(factory_code)
        
        # Flatten layout thành column_mapping động
        column_mapping = []
        for grp in config['layout']:
            for col in grp['columns']:
                device_code = grp['device_code']
                # Nếu có sub_device thì nối thêm vào device code đầy đủ
                if 'sub_device' in col:
                    device_code = f"{device_code}.{col['sub_device']}"
                column_mapping.append({
                    'name': col['ten'],
                    'device': device_code,
                    'field': col['ma'],
                    'unit': col['don_vi']
                })
                
        expected_cols = len(column_mapping)

        # Đọc file Excel: Template có 3 hàng header, dữ liệu bắt đầu từ hàng 4
        try:
            raw_df = pd.read_excel(
                file,
                header=None,
                skiprows=3,
                nrows=num_cycles,
                engine="openpyxl",
            )
            # Lấy từ cột thứ 2 trở đi (bỏ cột index 0 là STT/Thời gian)
            raw_col_count = len(raw_df.columns)
            first_col_is_helper = (
                raw_col_count > expected_cols
                and _looks_like_index_or_time_column(raw_df.iloc[:, 0], num_cycles)
            )

            if first_col_is_helper:
                # New template: column A is STT/time, parameters start at B.
                df = raw_df.iloc[:, 1:]
            else:
                # Legacy template: parameters start at A. Keeping column A
                # prevents a one-cell shift and an empty last parameter.
                df = raw_df
            
            # Reset column index về 0..N để tiện truy cập
            df.columns = range(len(df.columns))
            
            # Đảm bảo số lượng cột khớp với expected_cols bằng cách cắt hoặc bù thêm cột None
            actual_cols = len(df.columns)
            if actual_cols < expected_cols:
                for col_idx in range(actual_cols, expected_cols):
                    df[col_idx] = None
            elif actual_cols > expected_cols:
                df = df.iloc[:, :expected_cols]
                
        except Exception as e:
            return JsonResponse({'error': f'Lỗi đọc file Excel: {str(e)}'}, status=400)

        imported_count = 0

        # Lấy danh sách thiết bị cần thiết
        thiet_bi_map = {}
        scoped_thiet_bi = filter_queryset_by_factory(
            ThietBi.objects.all(),
            request.user,
            'nha_may',
            'string',
        )

        required_devices = set(mapping['device'] for mapping in column_mapping)

        for ma_thiet_bi in required_devices:
            lookup_code = ma_thiet_bi
            if factory_code != 'SH' and not ma_thiet_bi.startswith(f'{factory_code}.TB'):
                # Thay thế SH.TB hoặc VS.TB bằng factory_code hiện tại
                lookup_code = ma_thiet_bi.replace('SH.TB', f'{factory_code}.TB').replace('VS.TB', f'{factory_code}.TB')

            try:
                thiet_bi = scoped_thiet_bi.get(ma_day_du=lookup_code)
                thiet_bi_map[ma_thiet_bi] = thiet_bi
            except ThietBi.DoesNotExist:
                # Fallback to TPP parent if sub-device is not found
                if '.TPP.' in lookup_code:
                    fallback_code = lookup_code.split('.TPP.')[0] + '.TPP'
                    try:
                        thiet_bi = scoped_thiet_bi.get(ma_day_du=fallback_code)
                        thiet_bi_map[ma_thiet_bi] = thiet_bi
                    except ThietBi.DoesNotExist:
                        continue
                else:
                    continue

        if not thiet_bi_map:
            return JsonResponse({'error': 'Không tìm thấy thiết bị trong phạm vi nhà máy được phân quyền'}, status=403)

        # Lấy trước các bản ghi hiện có
        existing_records = ThongSoVanHanh.objects.filter(
            thiet_bi__in=thiet_bi_map.values(),
            ngay_nhap=import_date
        )
        existing_lookup = {
            (rec.thiet_bi_id, rec.ma_thong_so, rec.thoi_diem_nhap): rec
            for rec in existing_records
        }

        to_create = []
        to_update = []

        # Xử lý từng dòng (48 chu kỳ)
        for index, row in df.iterrows():
            try:
                ngay = import_date
                thoi_diem = time_cycles_list[index] if index < len(time_cycles_list) else f"{index:02d}:00"

                # Tạo timezone-aware datetime
                datetime_str = f"{ngay} {thoi_diem}:00"
                naive_datetime = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
                
                import pytz
                vietnam_tz = pytz.timezone('Asia/Ho_Chi_Minh')
                timezone_aware_datetime = vietnam_tz.localize(naive_datetime)

                for col_index, mapping_info in enumerate(column_mapping):
                    try:
                        value = row.iloc[col_index] if col_index < len(row) else None
                        device_code = mapping_info['device']
                        field_name = mapping_info['field']
                        unit = mapping_info['unit']
                        param_name = mapping_info['name']

                        if pd.isna(value) or value == '' or value is None or value == '-':
                            value = None

                        if device_code not in thiet_bi_map:
                            continue

                        thiet_bi = thiet_bi_map[device_code]
                        gia_tri_str = str(value) if value is not None else None
                        nha_may_val = get_user_factory_name(request.user) or thiet_bi.nha_may

                        lookup_key = (thiet_bi.id, field_name, timezone_aware_datetime)
                        if lookup_key in existing_lookup:
                            thong_so = existing_lookup[lookup_key]
                            if (thong_so.gia_tri != gia_tri_str or
                                thong_so.don_vi != unit or
                                thong_so.ten_thong_so != param_name or
                                thong_so.nha_may != nha_may_val):
                                
                                thong_so.gia_tri = gia_tri_str
                                thong_so.don_vi = unit
                                thong_so.ten_thong_so = param_name
                                thong_so.nha_may = nha_may_val
                                to_update.append(thong_so)
                        else:
                            thong_so = ThongSoVanHanh(
                                thiet_bi=thiet_bi,
                                ma_thong_so=field_name,
                                thoi_diem_nhap=timezone_aware_datetime,
                                gia_tri=gia_tri_str,
                                don_vi=unit,
                                ten_thong_so=param_name,
                                ngay_nhap=ngay,
                                nha_may=nha_may_val
                            )
                            to_create.append(thong_so)
                            existing_lookup[lookup_key] = thong_so

                        imported_count += 1

                    except Exception:
                        continue

            except Exception:
                continue

        # Ghi hàng loạt
        if to_create:
            ThongSoVanHanh.objects.bulk_create(to_create)
        if to_update:
            ThongSoVanHanh.objects.bulk_update(to_update, ['gia_tri', 'don_vi', 'ten_thong_so', 'nha_may'])

        return JsonResponse({
            'message': 'Import thành công',
            'imported_count': imported_count
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
