import io
import pandas as pd
from datetime import datetime, time
import pytz
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from rest_framework import viewsets, filters, status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from core.factory_scope import (
    filter_queryset_by_factory,
    get_user_factory_name,
    get_user_factory_code,
    apply_request_factory_to_serializer,
    has_all_factory_access,
    has_profile_permission,
)
from quanlyvanhanh.models import ThietBi, ThongSoTram110KV
from quanlyvanhanh.serializers import (
    ThongSoTram110KVSerializer,
    ThongSoTram110KVCreateSerializer,
)
from quanlyvanhanh.configs.operation_configs import TRAM_CONFIGS, get_tram_factory_config
from quanlyvanhanh.services.thongso_tram_service import (
    bulk_upsert_thong_so_tram_110kv,
)


def get_factory_config(factory_code):
    """Lay cau hinh tram 110kV theo ma nha may."""
    return get_tram_factory_config(factory_code)


def _ensure_thiet_bi_access(user, thiet_bi):
    if not thiet_bi:
        return
    allowed = filter_queryset_by_factory(
        ThietBi.objects.filter(pk=thiet_bi.pk),
        user,
        "nha_may",
        "string",
    ).exists()
    if not allowed:
        raise PermissionDenied("Bạn không có quyền thao tác với thiết bị này.")


class ThongSoTram110KVViewSet(viewsets.ModelViewSet):
    """ViewSet cho phép CRUD thông số trạm 110kV và các API động liên quan"""

    queryset = ThongSoTram110KV.objects.select_related("thiet_bi").all()
    serializer_class = ThongSoTram110KVSerializer

    def check_permissions(self, request):
        super().check_permissions(request)
        
        # Check custom profile permissions based on action
        if self.action in ["list", "retrieve", "by_date", "config"]:
            if not has_profile_permission(request.user, "can_view_operation_parameters"):
                self.permission_denied(
                    request,
                    message="Tài khoản của bạn chưa được cấp quyền xem thông số vận hành. Vui lòng liên hệ quản trị viên."
                )
        elif self.action in ["create", "update", "partial_update", "bulk_create", "bulk_upsert"]:
            if not has_profile_permission(request.user, "can_edit_operation_parameters"):
                self.permission_denied(
                    request,
                    message="Tài khoản của bạn chưa được cấp quyền thêm/sửa thông số vận hành. Vui lòng liên hệ quản trị viên."
                )
        elif self.action in ["destroy", "delete_by_day"]:
            if not has_profile_permission(request.user, "can_delete_operation_parameters"):
                self.permission_denied(
                    request,
                    message="Tài khoản của bạn chưa được cấp quyền xóa dữ liệu thông số vận hành. Vui lòng liên hệ quản trị viên."
                )
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["thiet_bi", "ten_thong_so", "nha_may", "ngay_nhap"]
    search_fields = ["ten_thong_so", "ma_thong_so", "thiet_bi__ten", "ghi_chu"]
    ordering_fields = ["ten_thong_so", "ngay_nhap", "thoi_diem_nhap", "created_at"]
    ordering = ["-ngay_nhap", "-thoi_diem_nhap"]

    def get_serializer_class(self):
        if self.action in ["create", "bulk_upsert", "bulk_create"]:
            return ThongSoTram110KVCreateSerializer
        return ThongSoTram110KVSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        return filter_queryset_by_factory(
            queryset,
            self.request.user,
            "nha_may",
            "string",
        )

    def perform_create(self, serializer):
        _ensure_thiet_bi_access(self.request.user, serializer.validated_data.get("thiet_bi"))
        serializer.save(
            **apply_request_factory_to_serializer(
                self.request.user,
                serializer,
                "nha_may",
                "string",
            )
        )

    def perform_update(self, serializer):
        _ensure_thiet_bi_access(
            self.request.user,
            serializer.validated_data.get("thiet_bi", serializer.instance.thiet_bi),
        )
        serializer.save(
            **apply_request_factory_to_serializer(
                self.request.user,
                serializer,
                "nha_may",
                "string",
            )
        )

    @action(detail=False, methods=["get"])
    def by_date(self, request):
        date_value = request.query_params.get("date") or request.query_params.get("ngay")
        if not date_value:
            return Response(
                {"error": "Cần cung cấp tham số date hoặc ngay"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        queryset = self.get_queryset().filter(ngay_nhap=date_value)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["post"])
    def bulk_create(self, request):
        return self.bulk_upsert(request)

    @action(detail=False, methods=["post"])
    def bulk_upsert(self, request):
        data_list = request.data
        if not isinstance(data_list, list):
            return Response(
                {"error": "Dữ liệu phải là một mảng"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            result = bulk_upsert_thong_so_tram_110kv(request.user, data_list)
        except PermissionDenied:
            raise
        except Exception as exc:
            return Response(
                {"error": f"Lỗi khi xử lý dữ liệu: {exc}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response(
            {
                **result,
                "message": (
                    f"Đã tạo {result['created']} bản ghi mới, "
                    f"cập nhật {result['updated']} bản ghi, "
                    f"xóa {result.get('deleted', 0)} bản ghi"
                ),
            },
            status=(
                status.HTTP_201_CREATED
                if result["created"] > 0
                else status.HTTP_200_OK
            ),
        )

    @action(detail=False, methods=["delete"])
    def delete_by_day(self, request):
        ngay_str = request.query_params.get("ngay")
        if not ngay_str:
            return Response(
                {"error": "Cần cung cấp tham số ngày (ngay)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        queryset = self.get_queryset().filter(ngay_nhap=ngay_str)
        deleted_count = queryset.delete()[0]
        return Response(
            {
                "message": f"Đã xóa {deleted_count} thông số trạm 110kV",
                "deleted_count": deleted_count,
            }
        )

    @action(detail=False, methods=["get"])
    def config(self, request):
        """API trả về cấu hình cột động theo nhà máy của tài khoản đang đăng nhập"""
        factory_code = get_user_factory_code(request.user) or 'SH'
        config = get_factory_config(factory_code)
        return Response({
            **config,
            "factory_code": factory_code,
            "nha_may": get_user_factory_name(request.user) or factory_code,
        })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def excel_template_tram(request):
    """Tải template Excel Thông số Trạm 110kV tự động sinh ra dựa theo cấu hình nhà máy"""
    try:
        if not has_profile_permission(request.user, "can_export_excel"):
            return JsonResponse({'error': 'Tài khoản của bạn chưa được cấp quyền xuất dữ liệu hoặc tải template Excel. Vui lòng liên hệ quản trị viên.'}, status=403)
        factory_code = get_user_factory_code(request.user) or 'SH'
        config = get_factory_config(factory_code)
        columns_config = config['columns']

        # Header 1: Tên bảng (Merger A1 đến hết số lượng cột)
        header_row1 = [config['title']] + [""] * (len(columns_config) - 1)
        # Header 2: Tên thông số
        header_row2 = [col["ten"] for col in columns_config]
        # Header 3: Đơn vị
        header_row3 = [col["don_vi"] for col in columns_config]

        data = [header_row1, header_row2, header_row3]
        for _ in range(12):
            # Khởi tạo mặc định dấu "-" cho các ô trống dữ liệu
            data.append(["-"] * len(columns_config))

        df = pd.DataFrame(data)

        # Ghi ra file Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_name = 'Thông số Trạm'
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Format Excel chuyên nghiệp
            # Merge tiêu đề dòng 1
            max_col_letter = chr(64 + len(columns_config)) if len(columns_config) <= 26 else f"A{chr(64 + len(columns_config) - 26)}"
            worksheet.merge_cells(f'A1:{max_col_letter}1')

            # Style A1
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet['A1'].fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid") # LightSteelBlue
            worksheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

            # Style Row 2 (Tên thông số)
            for col in range(1, len(columns_config) + 1):
                cell = worksheet.cell(row=2, column=col)
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")

            # Style Row 3 (Đơn vị)
            for col in range(1, len(columns_config) + 1):
                cell = worksheet.cell(row=3, column=col)
                cell.font = Font(italic=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")

            # Set độ rộng cột
            for col in range(1, len(columns_config) + 1):
                col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
                worksheet.column_dimensions[col_letter].width = 18

            # Kẻ viền cho bảng
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for r in range(1, 16):
                for col in range(1, len(columns_config) + 1):
                    worksheet.cell(row=r, column=col).border = thin_border

        file_content = output.getvalue()
        response = HttpResponse(
            file_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ThongSoTram_110kV_Template.xlsx"'
        return response

    except Exception as e:
        return HttpResponse(f'Lỗi khi sinh template Excel: {str(e)}', status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def excel_import_tram(request):
    """Import dữ liệu trạm 110kV từ file Excel và thực hiện mapping động dựa theo config"""
    try:
        if not has_profile_permission(request.user, "can_import_excel"):
            return JsonResponse({'error': 'Tài khoản của bạn chưa được cấp quyền import dữ liệu từ Excel. Vui lòng liên hệ quản trị viên.'}, status=403)
        excel_file = request.FILES.get('file')
        if not excel_file:
            return HttpResponse('Thiếu file tải lên', status=400)

        request_data = getattr(request, "data", request.POST)
        selected_date = request_data.get('selected_date')
        if not selected_date:
            return HttpResponse('Thiếu tham số selected_date', status=400)

        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse('Định dạng ngày không hợp lệ, sử dụng YYYY-MM-DD', status=400)

        # Lọc cấu hình nhà máy
        factory_code = get_user_factory_code(request.user) or 'SH'
        config = get_factory_config(factory_code)
        columns_config = config['columns']

        df = pd.read_excel(excel_file, header=None)

        # VALIDATION
        if len(df) < 3:
            return JsonResponse({
                'error': f'File Excel không hợp lệ: Cần ít nhất 3 hàng header. Hiện tại có {len(df)} hàng.',
                'status': 'error'
            }, status=400)

        # Xác thực số lượng cột (18 cột thông số xếp ngang từ A đến R)
        expected_cols = len(columns_config)
        if len(df.columns) < expected_cols:
            detected_headers = df.iloc[1].fillna("").astype(str).tolist() if len(df) > 1 else []
            return JsonResponse({
                'error': f'File Excel thiếu cột: Cần có ít nhất {expected_cols} cột thông số. Hiện tại file của bạn chỉ có {len(df.columns)} cột.',
                'status': 'error',
                'expected_columns': expected_cols,
                'actual_columns': len(df.columns),
                'detected_headers': detected_headers
            }, status=400)

        # Xác thực tiêu đề cột ở dòng 2 (Header Row 2)
        header_row2 = df.iloc[1, 0:expected_cols].fillna("").astype(str).tolist()
        missing_params = []
        for idx, col in enumerate(columns_config):
            actual_title = header_row2[idx].strip().lower()
            expected_title = col["ten"].strip().lower()
            # Chuẩn hóa loại bỏ toàn bộ khoảng trắng để so sánh không nhạy cảm khoảng trắng
            actual_norm = "".join(actual_title.split())
            expected_norm = "".join(expected_title.split())
            if expected_norm != actual_norm:
                missing_params.append(f'Cột {chr(65 + idx)}: Kỳ vọng "{col["ten"]}" nhưng nhận "{header_row2[idx]}"')

        if missing_params:
            return JsonResponse({
                'error': 'File Excel có tiêu đề cột không khớp mẫu trạm 110kV. Vui lòng tải file mẫu mới nhất.',
                'status': 'error',
                'missing_params': missing_params
            }, status=400)

        # Lấy trước danh sách thiết bị cần ánh xạ trong nhà máy của user
        allowed_thiet_bi = filter_queryset_by_factory(
            ThietBi.objects.all(),
            request.user,
            'nha_may',
            'string'
        )
        device_codes = set(col["ma_thiet_bi"] for col in columns_config)
        device_map = {
            tb.ma_day_du: tb for tb in allowed_thiet_bi.filter(ma_day_du__in=device_codes)
        }

        # Kiểm tra xem các thiết bị cấu hình đã tồn tại trong DB chưa
        for code in device_codes:
            if code not in device_map:
                # Tìm xem có thiết bị nào mà chưa khớp phân quyền không
                if ThietBi.objects.filter(ma_day_du=code).exists():
                    return JsonResponse({
                        'error': f'Bạn không có quyền thao tác với thiết bị {code} của nhà máy này.',
                        'status': 'error'
                    }, status=403)
                else:
                    return JsonResponse({
                        'error': f'Thiết bị {code} chưa được khai báo trên hệ thống.',
                        'status': 'error'
                    }, status=400)

        # Định dạng thời gian cho 12 chu kỳ (2h 1 lần: 0, 2, ..., 22)
        vietnam_tz = pytz.timezone('Asia/Ho_Chi_Minh')
        time_cycles = []
        for hour in range(0, 24, 2):
            naive_dt = datetime.combine(target_date, time(hour, 0))
            time_cycles.append(vietnam_tz.localize(naive_dt))

        # Thực hiện đọc và chuẩn bị dữ liệu ghi vào database
        upsert_payload = []
        for row_idx in range(3, min(15, len(df))): # Đọc 12 dòng dữ liệu bắt đầu từ dòng 4 đến dòng 15
            time_val = time_cycles[row_idx - 3]
            for col_idx, col_config in enumerate(columns_config, start=0):
                raw_val = df.iloc[row_idx, col_idx]
                
                # Check các ô mang dấu gạch ngang hoặc rỗng
                if pd.isna(raw_val) or raw_val == '' or raw_val == '-':
                    val_str = None
                else:
                    # Giữ nguyên giá trị (số hoặc chữ như BT) dạng string
                    val_str = str(raw_val).strip()

                upsert_payload.append({
                    "thiet_bi": device_map[col_config["ma_thiet_bi"]].id,
                    "ten_thong_so": col_config["ten"],
                    "ma_thong_so": col_config["ma"],
                    "don_vi": col_config["don_vi"],
                    "gia_tri": val_str,
                    "nha_may": factory_code,
                    "ky_hieu_van_hanh": "",
                    "thoi_diem_nhap": time_val.isoformat(),
                    "ngay_nhap": target_date.strftime('%Y-%m-%d')
                })

        # Lưu dữ liệu
        result = bulk_upsert_thong_so_tram_110kv(request.user, upsert_payload)
        return JsonResponse({
            'message': f"Import thành công: tạo {result['created']} bản ghi, cập nhật {result['updated']} bản ghi.",
            'status': 'success',
            'imported_count': result['created'] + result['updated']
        })

    except Exception as e:
        return HttpResponse(f'Lỗi khi import dữ liệu: {str(e)}', status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_thong_so_tram_110kv(request):
    """Xuất dữ liệu thông số trạm 110kV ra Excel theo layout chuẩn của nhà máy"""
    try:
        if not has_profile_permission(request.user, "can_export_excel"):
            return HttpResponse('Tài khoản của bạn chưa được cấp quyền xuất dữ liệu Excel. Vui lòng liên hệ quản trị viên.', status=403, content_type='text/plain; charset=utf-8')
        params = getattr(request, "query_params", request.GET)
        date = params.get('date')
        if not date:
            return HttpResponse('Thiếu tham số date', status=400)

        try:
            export_date = datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse('Định dạng ngày không hợp lệ, sử dụng YYYY-MM-DD', status=400)

        factory_code = get_user_factory_code(request.user) or 'SH'
        config = get_factory_config(factory_code)
        columns_config = config['columns']

        # Lấy dữ liệu từ DB
        queryset = filter_queryset_by_factory(
            ThongSoTram110KV.objects.select_related('thiet_bi').filter(ngay_nhap=export_date),
            request.user,
            "nha_may",
            "string",
        ).order_by('thoi_diem_nhap')

        # Map dữ liệu để đưa lên ô lưới Excel: { thoi_diem_gio: { ten_thong_so: gia_tri } }
        data_dict = {}
        for obj in queryset:
            hour_str = obj.thoi_diem_nhap.astimezone(pytz.timezone('Asia/Ho_Chi_Minh')).strftime('%H:%M')
            if hour_str not in data_dict:
                data_dict[hour_str] = {}
            data_dict[hour_str][obj.ten_thong_so] = obj.gia_tri

        # Tạo file Excel
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Thông số Trạm'

        # Dòng 1: Tiêu đề lớn
        max_col_letter = chr(64 + len(columns_config)) if len(columns_config) <= 26 else f"A{chr(64 + len(columns_config) - 26)}"
        ws.merge_cells(f'A1:{max_col_letter}1')
        ws['A1'] = config['title']
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        # Dòng 2: Tên thông số
        for col_idx, col_cfg in enumerate(columns_config, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_cfg["ten"])
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")

        # Dòng 3: Đơn vị
        for col_idx, col_cfg in enumerate(columns_config, start=1):
            cell = ws.cell(row=3, column=col_idx, value=col_cfg["don_vi"])
            cell.font = Font(italic=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")

        # Dòng 4 đến 15: 12 chu kỳ giờ (00:00, 02:00, ..., 22:00)
        for r_idx in range(4, 16):
            hour = (r_idx - 4) * 2
            hour_str = f"{hour:02d}:00"

            # Đổ dữ liệu thông số vào từng cột
            for col_idx, col_cfg in enumerate(columns_config, start=1):
                val = data_dict.get(hour_str, {}).get(col_cfg["ten"], None)
                ws.cell(row=r_idx, column=col_idx, value=val)

        # Format widths & borders
        for col in range(1, len(columns_config) + 1):
            col_letter = chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"
            ws.column_dimensions[col_letter].width = 18

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for r in range(1, 16):
            for col in range(1, len(columns_config) + 1):
                ws.cell(row=r, column=col).border = thin_border

        wb.save(buffer)
        buffer.seek(0)
        file_content = buffer.getvalue()

        response = HttpResponse(
            file_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="ThongSoTram_110kV_{date}.xlsx"'
        return response

    except Exception as e:
        return HttpResponse(f'Lỗi khi xuất dữ liệu: {str(e)}', status=500)
