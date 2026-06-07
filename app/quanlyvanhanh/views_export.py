import io
from django.http import HttpResponse
from datetime import datetime, timedelta, time
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from core.factory_scope import filter_queryset_by_factory, get_user_factory_code, has_profile_permission
from .models import ThongSoVanHanh, ThongSoToMay, ThietBi


def _query_params(request):
    return getattr(request, "query_params", request.GET)


def _scoped_thiet_bi_queryset(user):
    return filter_queryset_by_factory(
        ThietBi.objects.all(),
        user,
        "nha_may",
        "string",
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_thong_so(request):
    """Xuất dữ liệu thông số vận hành ra file Excel"""
    try:
        if not has_profile_permission(request.user, "can_export_excel"):
            return HttpResponse('Tài khoản của bạn chưa được cấp quyền xuất dữ liệu Excel. Vui lòng liên hệ quản trị viên.', status=403, content_type='text/plain; charset=utf-8')
        # Lấy tham số từ request
        params = _query_params(request)
        date = params.get('date')
        thiet_bi = params.get('thiet_bi', 'all')
        format_type = params.get('format', 'xlsx')


        if not date:
            return HttpResponse(
                'Thiếu tham số date',
                status=400
            )

        # Parse ngày
        try:
            export_date = datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse(
                'Định dạng ngày không hợp lệ, sử dụng YYYY-MM-DD',
                status=400
            )

        # Lấy dữ liệu thông số vận hành
        queryset = filter_queryset_by_factory(
            ThongSoVanHanh.objects.select_related('thiet_bi').filter(ngay_nhap=export_date),
            request.user,
            "nha_may",
            "string",
        )

        # Filter theo thiết bị nếu không phải "all"
        if thiet_bi != 'all':
            queryset = queryset.filter(thiet_bi__ma_day_du__startswith=thiet_bi)


        # Chuyển đổi thành DataFrame
        data = []
        for obj in queryset:
            data.append({
                'Ngày': obj.ngay_nhap,
                'Thời điểm': obj.thoi_diem_nhap.strftime('%H:%M'),
                'Thiết bị': obj.thiet_bi.ten,
                'Tên thông số': obj.ten_thong_so,
                'Giá trị': obj.gia_tri,
                'Đơn vị': obj.don_vi,
                'Nhà máy': obj.nha_may,
                'Ký hiệu vận hành': obj.ky_hieu_van_hanh,
                'Ghi chú': obj.ghi_chu,
            })

        if not data:
            return HttpResponse(
                'Không có dữ liệu để xuất cho ngày đã chọn',
                status=200,  # Trả về 200 thay vì 404 để frontend có thể handle
                content_type='text/plain; charset=utf-8'
            )


        # Tạo DataFrame
        df = pd.DataFrame(data)

        # Tạo file Excel
        if format_type == 'xlsx':
            # Sử dụng openpyxl trực tiếp (giống template)
            buffer = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = 'Thông số vận hành'

            # Ghi dữ liệu từ DataFrame vào worksheet
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Save workbook vào buffer (giống template)
            wb.save(buffer)
            buffer.seek(0)
            file_content = buffer.getvalue()

            # Tạo response
            response = HttpResponse(
                file_content,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="thong_so_van_hanh_{date}.xlsx"'
            response['Content-Length'] = len(file_content)
            response['Access-Control-Allow-Origin'] = '*'
            response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type'

        elif format_type == 'csv':
            csv_data = df.to_csv(index=False, encoding='utf-8-sig')
            response = HttpResponse(csv_data, content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="thong_so_van_hanh_{date}.csv"'
            response['Access-Control-Allow-Origin'] = '*'
            response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type'

        else:
            return HttpResponse(
                'Định dạng không được hỗ trợ',
                status=400
            )

        return response

    except Exception as e:
        return HttpResponse(
            f'Lỗi khi xuất dữ liệu: {str(e)}',
            status=500
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_thong_so_to_may_h1(request):
    """Xuất dữ liệu thông số tổ máy H1 ra file Excel"""
    try:
        if not has_profile_permission(request.user, "can_export_excel"):
            return HttpResponse('Tài khoản của bạn chưa được cấp quyền xuất dữ liệu Excel. Vui lòng liên hệ quản trị viên.', status=403, content_type='text/plain; charset=utf-8')
        # Lấy tham số từ request
        params = _query_params(request)
        date = params.get('date')
        format_type = params.get('format', 'xlsx')

        if not date:
            return HttpResponse(
                'Thiếu tham số date',
                status=400
            )

        # Parse ngày
        try:
            export_date = datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse(
                'Định dạng ngày không hợp lệ, sử dụng YYYY-MM-DD',
                status=400
            )

        # Lấy thiết bị H1
        factory_code = get_user_factory_code(request.user) or 'SH'
        try:
            thiet_bi = _scoped_thiet_bi_queryset(request.user).get(ma_day_du=f'{factory_code}.TB.H1.GE')
        except ThietBi.DoesNotExist:
            return HttpResponse(
                'Không tìm thấy thiết bị H1',
                status=404
            )

        # Lấy dữ liệu thông số tổ máy H1 (bao gồm cả các thiết bị con)
        prefix = ".".join(thiet_bi.ma_day_du.split(".")[:3])  # E.g., "SH.TB.H1"
        queryset = filter_queryset_by_factory(
            ThongSoToMay.objects.select_related('thiet_bi').filter(
                thiet_bi__ma_day_du__startswith=prefix,
                ngay_nhap=export_date
            ),
            request.user,
            "nha_may",
            "string",
        ).order_by('thoi_diem_nhap', 'ten_thong_so')

        if not queryset.exists():
            return HttpResponse(
                'Không có dữ liệu để xuất cho ngày đã chọn',
                status=200,
                content_type='text/plain; charset=utf-8'
            )

        # Tạo file Excel với format giống template
        if format_type == 'xlsx':
            buffer = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = 'Thông số H1'

            # Định nghĩa các thông số (giống template)
            thong_so_list = [
                {"ten": "Áp lực nước", "ma": "ap_luc_nuoc", "don_vi": "bar"},
                {"ten": "Áp lực chèn trục", "ma": "ap_luc_chen_truc", "don_vi": "bar"},
                {"ten": "Lưu lượng chèn trục", "ma": "luu_luong_chen_truc", "don_vi": "l/p"},
                {"ten": "Lưu lượng ổ hướng tuabin", "ma": "luu_luong_o_huong_tuabin", "don_vi": "l/p"},
                {"ten": "Nhiệt độ ổ hướng tuabin", "ma": "nhiet_do_o_huong_tuabin", "don_vi": "°C"},
                {"ten": "Lưu lượng ổ hướng máy phát", "ma": "luu_luong_o_huong_may_phat", "don_vi": "l/p"},
                {"ten": "Nhiệt độ ổ hướng máy phát", "ma": "nhiet_do_o_huong_may_phat", "don_vi": "°C"},
                {"ten": "Lưu lượng ổ đỡ máy phát", "ma": "luu_luong_o_do_may_phat", "don_vi": "l/p"},
                {"ten": "Nhiệt độ ổ đỡ", "ma": "nhiet_do_o_do", "don_vi": "°C"},
                {"ten": "Nhiệt độ ổ hướng - ổ đỡ", "ma": "nhiet_do_o_huong_o_do", "don_vi": "°C"},
                {"ten": "Nhiệt độ đầu ổ đỡ", "ma": "nhiet_do_dau_o_do", "don_vi": "°C"},
                {"ten": "Lưu lượng làm mát máy phát", "ma": "luu_luong_lam_mat_may_phat", "don_vi": "l/p"},
                {"ten": "Nhiệt độ nước làm mát máy phát", "ma": "nhiet_do_nuoc_lam_mat_may_phat", "don_vi": "°C"},
                {"ten": "Nhiệt độ khí mát", "ma": "nhiet_do_khi_mat", "don_vi": "°C"},
                {"ten": "Nhiệt độ khí nóng", "ma": "nhiet_do_khi_nong", "don_vi": "°C"},
                {"ten": "Nhiệt độ cuộn dây stato", "ma": "nhiet_do_cuon_day_stato", "don_vi": "°C"},
                {"ten": "Tốc độ", "ma": "toc_do", "don_vi": "v/ph"},
                {"ten": "Giới hạn độ mở cánh hướng", "ma": "gioi_han_do_mo_canh_huong", "don_vi": "%"},
                {"ten": "Độ mở cánh hướng", "ma": "do_mo_canh_huong", "don_vi": "%"},
                {"ten": "Độ rơi tốc", "ma": "do_roi_toc", "don_vi": "%"},
            ]

            # Header row 1: "THÔNG SỐ H1"
            ws.merge_cells('A1:T1')
            ws['A1'] = 'THÔNG SỐ H1'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

            # Header row 2: Tên thông số
            for col_idx, ts in enumerate(thong_so_list, start=1):
                cell = ws.cell(row=2, column=col_idx, value=ts["ten"])
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Header row 3: Đơn vị
            for col_idx, ts in enumerate(thong_so_list, start=1):
                cell = ws.cell(row=3, column=col_idx, value=ts["don_vi"])
                cell.font = Font(size=10)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data rows: 24 giờ (00:00 - 23:00)
            time_slots = [time(hour, 0) for hour in range(24)]
            data_dict = {}  # {thoi_diem: {ma_thong_so: gia_tri}}

            for obj in queryset:
                thoi_diem = obj.thoi_diem_nhap.time() if hasattr(obj.thoi_diem_nhap, 'time') else obj.thoi_diem_nhap
                if thoi_diem not in data_dict:
                    data_dict[thoi_diem] = {}
                data_dict[thoi_diem][obj.ten_thong_so] = obj.gia_tri

            for row_idx, time_slot in enumerate(time_slots, start=4):
                # Time column (A)
                time_str = time_slot.strftime('%H:%M')
                ws.cell(row=row_idx, column=1, value=time_str)
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
                ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")

                # Data columns (B-T)
                for col_idx, ts in enumerate(thong_so_list, start=2):
                    gia_tri = data_dict.get(time_slot, {}).get(ts["ten"], None)
                    ws.cell(row=row_idx, column=col_idx, value=gia_tri)

            # Set column widths
            ws.column_dimensions['A'].width = 12
            for col in range(2, len(thong_so_list) + 2):
                ws.column_dimensions[chr(64 + col)].width = 15

            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in range(1, 28):
                for col in range(1, len(thong_so_list) + 2):
                    ws.cell(row=row, column=col).border = thin_border

            wb.save(buffer)
            buffer.seek(0)
            file_content = buffer.getvalue()

            response = HttpResponse(
                file_content,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="thong_so_to_may_h1_{date}.xlsx"'
            response['Content-Length'] = len(file_content)
            response['Access-Control-Allow-Origin'] = '*'
            response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type'

            return response
        else:
            return HttpResponse(
                'Định dạng không được hỗ trợ',
                status=400
            )

    except Exception as e:
        return HttpResponse(
            f'Lỗi khi xuất dữ liệu H1: {str(e)}',
            status=500
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_thong_so_to_may_h2(request):
    """Xuất dữ liệu thông số tổ máy H2 ra file Excel"""
    try:
        if not has_profile_permission(request.user, "can_export_excel"):
            return HttpResponse('Tài khoản của bạn chưa được cấp quyền xuất dữ liệu Excel. Vui lòng liên hệ quản trị viên.', status=403, content_type='text/plain; charset=utf-8')
        # Lấy tham số từ request
        params = _query_params(request)
        date = params.get('date')
        format_type = params.get('format', 'xlsx')

        if not date:
            return HttpResponse(
                'Thiếu tham số date',
                status=400
            )

        # Parse ngày
        try:
            export_date = datetime.strptime(date, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse(
                'Định dạng ngày không hợp lệ, sử dụng YYYY-MM-DD',
                status=400
            )

        # Lấy thiết bị H2
        factory_code = get_user_factory_code(request.user) or 'SH'
        try:
            thiet_bi = _scoped_thiet_bi_queryset(request.user).get(ma_day_du=f'{factory_code}.TB.H2.GE')
        except ThietBi.DoesNotExist:
            return HttpResponse(
                'Không tìm thấy thiết bị H2',
                status=404
            )

        # Lấy dữ liệu thông số tổ máy H2 (bao gồm cả các thiết bị con)
        prefix = ".".join(thiet_bi.ma_day_du.split(".")[:3])  # E.g., "SH.TB.H2"
        queryset = filter_queryset_by_factory(
            ThongSoToMay.objects.select_related('thiet_bi').filter(
                thiet_bi__ma_day_du__startswith=prefix,
                ngay_nhap=export_date
            ),
            request.user,
            "nha_may",
            "string",
        ).order_by('thoi_diem_nhap', 'ten_thong_so')

        if not queryset.exists():
            return HttpResponse(
                'Không có dữ liệu để xuất cho ngày đã chọn',
                status=200,
                content_type='text/plain; charset=utf-8'
            )

        # Tạo file Excel với format giống template (giống H1 nhưng đổi H1 -> H2)
        if format_type == 'xlsx':
            buffer = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = 'Thông số H2'

            # Định nghĩa các thông số (giống H1)
            thong_so_list = [
                {"ten": "Áp lực nước", "ma": "ap_luc_nuoc", "don_vi": "bar"},
                {"ten": "Áp lực chèn trục", "ma": "ap_luc_chen_truc", "don_vi": "bar"},
                {"ten": "Lưu lượng chèn trục", "ma": "luu_luong_chen_truc", "don_vi": "l/p"},
                {"ten": "Lưu lượng ổ hướng tuabin", "ma": "luu_luong_o_huong_tuabin", "don_vi": "l/p"},
                {"ten": "Nhiệt độ ổ hướng tuabin", "ma": "nhiet_do_o_huong_tuabin", "don_vi": "°C"},
                {"ten": "Lưu lượng ổ hướng máy phát", "ma": "luu_luong_o_huong_may_phat", "don_vi": "l/p"},
                {"ten": "Nhiệt độ ổ hướng máy phát", "ma": "nhiet_do_o_huong_may_phat", "don_vi": "°C"},
                {"ten": "Lưu lượng ổ đỡ máy phát", "ma": "luu_luong_o_do_may_phat", "don_vi": "l/p"},
                {"ten": "Nhiệt độ ổ đỡ", "ma": "nhiet_do_o_do", "don_vi": "°C"},
                {"ten": "Nhiệt độ ổ hướng - ổ đỡ", "ma": "nhiet_do_o_huong_o_do", "don_vi": "°C"},
                {"ten": "Nhiệt độ đầu ổ đỡ", "ma": "nhiet_do_dau_o_do", "don_vi": "°C"},
                {"ten": "Lưu lượng làm mát máy phát", "ma": "luu_luong_lam_mat_may_phat", "don_vi": "l/p"},
                {"ten": "Nhiệt độ nước làm mát máy phát", "ma": "nhiet_do_nuoc_lam_mat_may_phat", "don_vi": "°C"},
                {"ten": "Nhiệt độ khí mát", "ma": "nhiet_do_khi_mat", "don_vi": "°C"},
                {"ten": "Nhiệt độ khí nóng", "ma": "nhiet_do_khi_nong", "don_vi": "°C"},
                {"ten": "Nhiệt độ cuộn dây stato", "ma": "nhiet_do_cuon_day_stato", "don_vi": "°C"},
                {"ten": "Tốc độ", "ma": "toc_do", "don_vi": "v/ph"},
                {"ten": "Giới hạn độ mở cánh hướng", "ma": "gioi_han_do_mo_canh_huong", "don_vi": "%"},
                {"ten": "Độ mở cánh hướng", "ma": "do_mo_canh_huong", "don_vi": "%"},
                {"ten": "Độ rơi tốc", "ma": "do_roi_toc", "don_vi": "%"},
            ]

            # Header row 1: "THÔNG SỐ H2"
            ws.merge_cells('A1:T1')
            ws['A1'] = 'THÔNG SỐ H2'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

            # Header row 2: Tên thông số
            for col_idx, ts in enumerate(thong_so_list, start=1):
                cell = ws.cell(row=2, column=col_idx, value=ts["ten"])
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Header row 3: Đơn vị
            for col_idx, ts in enumerate(thong_so_list, start=1):
                cell = ws.cell(row=3, column=col_idx, value=ts["don_vi"])
                cell.font = Font(size=10)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data rows: 24 giờ (00:00 - 23:00)
            time_slots = [time(hour, 0) for hour in range(24)]
            data_dict = {}  # {thoi_diem: {ma_thong_so: gia_tri}}

            for obj in queryset:
                thoi_diem = obj.thoi_diem_nhap.time() if hasattr(obj.thoi_diem_nhap, 'time') else obj.thoi_diem_nhap
                if thoi_diem not in data_dict:
                    data_dict[thoi_diem] = {}
                data_dict[thoi_diem][obj.ten_thong_so] = obj.gia_tri

            for row_idx, time_slot in enumerate(time_slots, start=4):
                # Time column (A)
                time_str = time_slot.strftime('%H:%M')
                ws.cell(row=row_idx, column=1, value=time_str)
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
                ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")

                # Data columns (B-T)
                for col_idx, ts in enumerate(thong_so_list, start=2):
                    gia_tri = data_dict.get(time_slot, {}).get(ts["ten"], None)
                    ws.cell(row=row_idx, column=col_idx, value=gia_tri)

            # Set column widths
            ws.column_dimensions['A'].width = 12
            for col in range(2, len(thong_so_list) + 2):
                ws.column_dimensions[chr(64 + col)].width = 15

            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in range(1, 28):
                for col in range(1, len(thong_so_list) + 2):
                    ws.cell(row=row, column=col).border = thin_border

            wb.save(buffer)
            buffer.seek(0)
            file_content = buffer.getvalue()

            response = HttpResponse(
                file_content,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="thong_so_to_may_h2_{date}.xlsx"'
            response['Content-Length'] = len(file_content)
            response['Access-Control-Allow-Origin'] = '*'
            response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
            response['Access-Control-Allow-Headers'] = 'Content-Type'

            return response
        else:
            return HttpResponse(
                'Định dạng không được hỗ trợ',
                status=400
            )

    except Exception as e:
        return HttpResponse(
            f'Lỗi khi xuất dữ liệu H2: {str(e)}',
            status=500
        )
