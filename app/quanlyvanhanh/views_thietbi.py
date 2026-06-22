import io
from datetime import datetime

import qrcode
from django.db.models import Q
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response

from core.factory_scope import (
    apply_request_factory_to_serializer,
    filter_queryset_by_factory,
    has_all_factory_access,
    has_profile_permission,
    get_user_factory_code,
)
from quanlyvanhanh.models import ThietBi
from quanlyvanhanh.serializers import (
    ThietBiDetailSerializer,
    ThietBiListSerializer,
    ThietBiSerializer,
)


class ThietBiPageNumberPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "limit"
    max_page_size = 200


def _ensure_thiet_bi_access(user, thiet_bi):
    if not thiet_bi or has_all_factory_access(user):
        return
    allowed = filter_queryset_by_factory(
        ThietBi.objects.filter(pk=thiet_bi.pk),
        user,
        "nha_may",
        "string",
    ).exists()
    if not allowed:
        raise PermissionDenied(
            "Ban khong co quyen thao tac voi thiet bi cua nha may nay."
        )


class ThietBiViewSet(viewsets.ModelViewSet):
    """ViewSet quan ly thiet bi."""

    queryset = ThietBi.objects.all()
    serializer_class = ThietBiSerializer
    pagination_class = ThietBiPageNumberPagination
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    filterset_fields = ["loai", "trang_thai", "nha_che_tao", "nha_cung_cap", "cap", "cha"]
    search_fields = [
        "ten",
        "ma",
        "ma_day_du",
        "ma_van_hanh",
        "so_serial",
        "mo_ta_ky_thuat",
    ]
    ordering_fields = ["ten", "ma_day_du", "thu_tu", "do_uu_tien", "cap"]
    ordering = ["cha__id", "thu_tu", "ten"]

    def check_permissions(self, request):
        super().check_permissions(request)

        if self.action in [
            "list",
            "retrieve",
            "con",
            "qr",
            "qr_export_items",
            "cay_phan_cap",
            "tim_kiem",
            "cap_0_codes",
            "cap_1_by_parent",
            "cap_2_by_parent",
            "excel_template",
            "export_excel",
        ]:
            if not has_profile_permission(request.user, "can_view_equipment"):
                self.permission_denied(
                    request,
                    message="Tài khoản của bạn chưa được cấp quyền xem danh sách thiết bị vận hành.",
                )
        elif self.action in ["create", "import_excel"]:
            if not has_profile_permission(request.user, "can_create_equipment"):
                self.permission_denied(
                    request,
                    message="Tài khoản của bạn chưa được cấp quyền tạo thiết bị vận hành.",
                )
        elif self.action in ["update", "partial_update"]:
            if not has_profile_permission(request.user, "can_edit_equipment"):
                self.permission_denied(
                    request,
                    message="Tài khoản của bạn chưa được cấp quyền sửa thiết bị vận hành.",
                )
        elif self.action == "destroy":
            if not has_profile_permission(request.user, "can_delete_equipment"):
                self.permission_denied(
                    request,
                    message="Tài khoản của bạn chưa được cấp quyền xóa thiết bị vận hành.",
                )

    def get_serializer_class(self):
        if self.action == "list":
            return ThietBiListSerializer
        if self.action == "retrieve":
            return ThietBiDetailSerializer
        return ThietBiSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = filter_queryset_by_factory(
            queryset,
            self.request.user,
            "nha_may",
            "string",
        )

        factory_param = (
            self.request.query_params.get("nha_may")
            or self.request.query_params.get("ma_nha_may")
        )
        if factory_param and str(factory_param).lower() != "all":
            factory_value = str(factory_param).strip()
            factory_query = (
                Q(nha_may__iexact=factory_value)
                | Q(ma_day_du__istartswith=f"{factory_value}.")
            )

            try:
                from khovattu.models import Bang_nha_may

                factory = Bang_nha_may.objects.filter(
                    Q(ma_nha_may__iexact=factory_value)
                    | Q(ten_nha_may__iexact=factory_value)
                ).first()
                if factory:
                    factory_query |= Q(nha_may__iexact=factory.ma_nha_may)
                    factory_query |= Q(nha_may__iexact=factory.ten_nha_may)
                    factory_query |= Q(nha_may__icontains=factory.ten_nha_may)
                    factory_query |= Q(ma_day_du__istartswith=f"{factory.ma_nha_may}.")
            except Exception:
                pass

            queryset = queryset.filter(factory_query)

        search_param = self.request.query_params.get("q")
        if search_param:
            search_param = str(search_param).strip()
            parts = [part for part in search_param.split(".") if part]

            if len(parts) >= 3 and all(part.isalnum() for part in parts[:3]):
                queryset = queryset.filter(
                    Q(ma_day_du__istartswith=search_param)
                    | Q(ma_day_du__iexact=search_param)
                )
            else:
                tokens = [token for token in search_param.replace(".", " ").split() if token]
                for token in tokens:
                    token_query = (
                        Q(ten__unaccent__icontains=token)
                        | Q(cha__ten__unaccent__icontains=token)
                        | Q(cha__cha__ten__unaccent__icontains=token)
                        | Q(cha__cha__cha__ten__unaccent__icontains=token)
                        | Q(cha__cha__cha__cha__ten__unaccent__icontains=token)
                        | Q(ma__unaccent__icontains=token)
                        | Q(ma_day_du__icontains=token)
                        | Q(ma_van_hanh__unaccent__icontains=token)
                        | Q(so_serial__unaccent__icontains=token)
                        | Q(mo_ta_ky_thuat__unaccent__icontains=token)
                    )
                    queryset = queryset.filter(token_query)

        return queryset

    def perform_create(self, serializer):
        _ensure_thiet_bi_access(self.request.user, serializer.validated_data.get("cha"))
        serializer.save(
            **apply_request_factory_to_serializer(
                self.request.user,
                serializer,
                "nha_may",
                "string",
            )
        )

    def perform_update(self, serializer):
        _ensure_thiet_bi_access(
            self.request.user,
            serializer.validated_data.get("cha", serializer.instance.cha),
        )
        serializer.save(
            **apply_request_factory_to_serializer(
                self.request.user,
                serializer,
                "nha_may",
                "string",
            )
        )

    @action(detail=True, methods=["get"])
    def con(self, request, pk=None):
        thiet_bi = self.get_object()
        con = thiet_bi.con.all()
        serializer = ThietBiListSerializer(con, many=True, context={"request": request})
        return Response(serializer.data)

    @action(detail=True, methods=["get"])
    def qr(self, request, pk=None):
        thiet_bi = self.get_object()
        from django.conf import settings
        frontend_base = None
        referer = request.META.get('HTTP_REFERER')
        if referer:
            from urllib.parse import urlparse
            parsed_uri = urlparse(referer)
            frontend_base = f"{parsed_uri.scheme}://{parsed_uri.netloc}"
            
        if not frontend_base:
            frontend_base = getattr(settings, 'KHO_QR_FRONTEND_BASE', 'http://localhost:5173')
            
        qr_data = f"{frontend_base}/quanlyvanhanh/thietbi?detailId={thiet_bi.pk}"
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=2,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")

        buffer = io.BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="image/png")
        response["Content-Disposition"] = (
            f'inline; filename="QR_{thiet_bi.ma_day_du.replace(".", "_")}.png"'
        )
        return response

    @action(detail=False, methods=["get"])
    def qr_export_items(self, request):
        queryset = self.filter_queryset(self.get_queryset()).order_by("ma_day_du")
        serializer = ThietBiListSerializer(queryset, many=True, context={"request": request})
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def cay_phan_cap(self, request):
        goc = self.get_queryset().filter(cha__isnull=True).order_by("thu_tu", "ten")
        serializer = ThietBiDetailSerializer(goc, many=True, context={"request": request})
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def tim_kiem(self, request):
        query = request.query_params.get("q", "")
        if query:
            queryset = self.get_queryset().filter(
                Q(ten__icontains=query)
                | Q(ma__icontains=query)
                | Q(ma_day_du__icontains=query)
                | Q(so_serial__icontains=query)
            )
        else:
            queryset = self.get_queryset()

        serializer = ThietBiListSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"])
    def cap_0_codes(self, request):
        try:
            all_thiet_bi = self.get_queryset().filter(
                ma_day_du__isnull=False
            ).values("ma_day_du", "ten")
            cap_0_info = {}

            for tb in all_thiet_bi:
                ma_day_du = tb["ma_day_du"]
                if ma_day_du:
                    parts = ma_day_du.split(".")
                    if len(parts) >= 3:
                        cap_0_code = ".".join(parts[:3])

                        if ma_day_du == cap_0_code:
                            cap_0_info[cap_0_code] = {
                                "ten": tb["ten"],
                                "ma": parts[-1],
                            }
                        elif cap_0_code not in cap_0_info:
                            cap_0_info[cap_0_code] = {
                                "ten": f"Thiet bi {parts[-1]}",
                                "ma": parts[-1],
                            }

            cap_0_devices = []
            for code in sorted(cap_0_info.keys()):
                info = cap_0_info[code]
                cap_0_devices.append({
                    "id": code,
                    "ma": info["ma"],
                    "ten": info["ten"],
                    "ma_day_du": code,
                    "cap": 0,
                })

            return Response(cap_0_devices)
        except Exception as exc:
            return Response(
                {"error": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"])
    def cap_1_by_parent(self, request):
        try:
            parent_code = request.query_params.get("parent_code", "")
            if not parent_code:
                return Response([])

            all_thiet_bi = self.get_queryset().filter(
                ma_day_du__startswith=parent_code + "."
            ).values("ma_day_du", "ten", "cap")
            cap_1_info = {}

            for tb in all_thiet_bi:
                ma_day_du = tb["ma_day_du"]
                if ma_day_du and ma_day_du.startswith(parent_code + "."):
                    parts = ma_day_du.split(".")
                    if len(parts) >= 4:
                        cap_1_code = ".".join(parts[:4])

                        if ma_day_du == cap_1_code:
                            cap_1_info[cap_1_code] = {
                                "ten": tb["ten"],
                                "ma": parts[3],
                            }
                        elif cap_1_code not in cap_1_info:
                            cap_1_info[cap_1_code] = {
                                "ten": f"Thiet bi {parts[3]}",
                                "ma": parts[3],
                            }

            cap_1_devices = []
            for code in sorted(cap_1_info.keys()):
                info = cap_1_info[code]
                cap_1_devices.append({
                    "id": code,
                    "ma": info["ma"],
                    "ten": info["ten"],
                    "ma_day_du": code,
                    "cap": 1,
                })

            return Response(cap_1_devices)
        except Exception as exc:
            return Response(
                {"error": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"])
    def cap_2_by_parent(self, request):
        try:
            parent_code = request.query_params.get("parent_code", "")
            if not parent_code:
                return Response([])

            all_thiet_bi = self.get_queryset().filter(
                ma_day_du__startswith=parent_code + "."
            ).values("ma_day_du", "ten", "cap")
            cap_2_info = {}

            for tb in all_thiet_bi:
                ma_day_du = tb["ma_day_du"]
                if ma_day_du and ma_day_du.startswith(parent_code + "."):
                    parts = ma_day_du.split(".")
                    if len(parts) >= 5:
                        cap_2_code = ".".join(parts[:5])

                        if ma_day_du == cap_2_code:
                            cap_2_info[cap_2_code] = {
                                "ten": tb["ten"],
                                "ma": parts[4],
                            }
                        elif cap_2_code not in cap_2_info:
                            cap_2_info[cap_2_code] = {
                                "ten": f"Thiet bi {parts[4]}",
                                "ma": parts[4],
                            }

            cap_2_devices = []
            for code in sorted(cap_2_info.keys()):
                info = cap_2_info[code]
                cap_2_devices.append({
                    "id": code,
                    "ma": info["ma"],
                    "ten": info["ten"],
                    "ma_day_du": code,
                    "cap": 2,
                })

            return Response(cap_2_devices)
        except Exception as exc:
            return Response(
                {"error": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=False, methods=["get"])
    def excel_template(self, request):
        """Tạo file template Excel để nhập danh sách thiết bị"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template Thiet Bi"

        headers = [
            "Mã thiết bị (*)",
            "Mã đầy đủ thiết bị chi tiết cha",
            "Tên thiết bị (*)",
            "Loại/Phân loại",
            "Trạng thái",
            "Nhà chế tạo",
            "Nhà cung cấp",
            "Nước sản xuất",
            "Nhà máy (*)",
            "Mã vận hành",
            "Bộ phận quản lý",
            "Bảng vẽ",
            "Thông số kỹ thuật",
            "Ngày lắp đặt (YYYY-MM-DD)",
            "Ngày vận hành (YYYY-MM-DD)"
        ]

        # Style header
        header_font = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color="BFBFBF"),
            right=Side(style='thin', color="BFBFBF"),
            top=Side(style='thin', color="BFBFBF"),
            bottom=Side(style='thin', color="BFBFBF")
        )

        ws.row_dimensions[1].height = 28

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = 25

        # Add sample row
        sample_row = [
            "PD.01",
            "SH.TB.H1.GE.OD",
            "Phân phối dầu tổ máy 01",
            "Thiết bị phụ",
            "Hoạt động",
            "Alstom",
            "Alstom Vietnam",
            "Pháp",
            "SH",
            "1-OPD",
            "Phân xưởng vận hành",
            "SH-H1-GE-01",
            "Áp lực định mức 40 bar",
            "2026-01-15",
            "2026-02-01"
        ]

        ws.row_dimensions[2].height = 20
        for col_idx, val in enumerate(sample_row, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.font = Font(name="Times New Roman", size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="ThietBi_Import_Template.xlsx"'
        return response

    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        """Xuất danh sách thiết bị hiện tại ra file Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # Lấy danh sách thiết bị theo bộ lọc hiện tại (dùng filter_queryset)
        queryset = self.filter_queryset(self.get_queryset()).order_by("ma_day_du")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sach Thiet Bi"

        headers = [
            "Mã thiết bị (*)",
            "Mã đầy đủ thiết bị chi tiết cha",
            "Tên thiết bị (*)",
            "Loại/Phân loại",
            "Trạng thái",
            "Nhà chế tạo",
            "Nhà cung cấp",
            "Nước sản xuất",
            "Nhà máy (*)",
            "Mã vận hành",
            "Bộ phận quản lý",
            "Bảng vẽ",
            "Thông số kỹ thuật",
            "Ngày lắp đặt (YYYY-MM-DD)",
            "Ngày vận hành (YYYY-MM-DD)"
        ]

        # Style header
        header_font = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color="BFBFBF"),
            right=Side(style='thin', color="BFBFBF"),
            top=Side(style='thin', color="BFBFBF"),
            bottom=Side(style='thin', color="BFBFBF")
        )

        ws.row_dimensions[1].height = 28

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = 25

        # Điền dữ liệu
        for row_idx, tb in enumerate(queryset, 2):
            ws.row_dimensions[row_idx].height = 20
            row_data = [
                tb.ma,
                tb.cha.ma_day_du if tb.cha else "",
                tb.ten,
                tb.loai or "",
                tb.trang_thai or "",
                tb.nha_che_tao or "",
                tb.nha_cung_cap or "",
                tb.nuoc_san_xuat or "",
                tb.nha_may or "",
                tb.ma_van_hanh or "",
                tb.bo_phan_quan_ly or "",
                tb.bang_ve or "",
                tb.mo_ta_ky_thuat or "",
                tb.ngay_lap_dat.strftime("%Y-%m-%d") if tb.ngay_lap_dat else "",
                tb.ngay_dua_vao_van_hanh.strftime("%Y-%m-%d") if tb.ngay_dua_vao_van_hanh else ""
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Times New Roman", size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="DanhSach_ThietBi.xlsx"'
        return response

    @action(detail=False, methods=["post"])
    def import_excel(self, request):
        """Import danh sách thiết bị từ file Excel"""
        import pandas as pd
        from django.db import transaction

        if 'file' not in request.FILES:
            return Response({'error': 'Không tìm thấy file upload.'}, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        try:
            df = pd.read_excel(file, engine="openpyxl")
        except Exception as e:
            return Response({'error': f'Lỗi đọc file Excel: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        # Chuẩn hóa tên cột
        df.columns = [str(col).strip() for col in df.columns]

        # Ánh xạ tên cột từ cũ/legacy/custom sang template chuẩn
        col_mappings = {
            "Mã thiết bị (*)": ["Mã thiết bị", "Mã cấp hiện tại", "Mã"],
            "Mã đầy đủ thiết bị chi tiết cha": ["Mã đầy đủ thiết bị cha"],
            "Tên thiết bị (*)": ["Tên thiết bị"],
            "Nhà máy (*)": ["Nhà máy"],
            "Ngày lắp đặt (YYYY-MM-DD)": ["Ngày lắp đặt"],
            "Ngày vận hành (YYYY-MM-DD)": ["Ngày vận hành"],
        }

        for target_col, source_cols in col_mappings.items():
            if target_col not in df.columns:
                for src in source_cols:
                    if src in df.columns:
                        df = df.rename(columns={src: target_col})
                        break

        # Kiểm tra các cột bắt buộc (phải có Mã thiết bị (*) hoặc Mã đầy đủ, Tên thiết bị (*), Nhà máy (*))
        required_cols = []
        if "Mã thiết bị (*)" not in df.columns and "Mã đầy đủ" not in df.columns:
            required_cols.append("Mã thiết bị (*)")
        if "Tên thiết bị (*)" not in df.columns:
            required_cols.append("Tên thiết bị (*)")
        if "Nhà máy (*)" not in df.columns:
            required_cols.append("Nhà máy (*)")

        if required_cols:
            return Response(
                {'error': f'File Excel thiếu các cột bắt buộc: {", ".join(required_cols)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        user_factory = get_user_factory_code(request.user)
        has_all_access = has_all_factory_access(request.user)

        success_count = 0
        errors = []

        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # Lấy mã thiết bị, nếu trống thử lấy từ Mã đầy đủ
                    ma = str(row.get("Mã thiết bị (*)")).strip() if pd.notna(row.get("Mã thiết bị (*)")) else ""
                    if not ma:
                        ma_day_du_col = row.get("Mã đầy đủ")
                        if pd.notna(ma_day_du_col) and str(ma_day_du_col).strip():
                            ma_day_du_str = str(ma_day_du_col).strip()
                            cha_ma_day_du = row.get("Mã đầy đủ thiết bị chi tiết cha")
                            if pd.notna(cha_ma_day_du) and str(cha_ma_day_du).strip():
                                parent_str = str(cha_ma_day_du).strip()
                                if ma_day_du_str.startswith(parent_str + "."):
                                    ma = ma_day_du_str[len(parent_str) + 1:]
                                else:
                                    parts = ma_day_du_str.split(".")
                                    ma = parts[-1]
                            else:
                                # Tự động phân tách cha và con từ Mã đầy đủ
                                parts = ma_day_du_str.split(".")
                                if len(parts) > 1:
                                    found_parent = None
                                    for i in range(len(parts) - 1, 0, -1):
                                        prefix = ".".join(parts[:i])
                                        if ThietBi.objects.filter(ma_day_du=prefix).exists():
                                            found_parent = prefix
                                            ma = ".".join(parts[i:])
                                            break
                                    if found_parent:
                                        row["Mã đầy đủ thiết bị chi tiết cha"] = found_parent
                                    else:
                                        ma = ma_day_du_str
                                else:
                                    ma = ma_day_du_str

                    ten = str(row.get("Tên thiết bị (*)")).strip() if pd.notna(row.get("Tên thiết bị (*)")) else ""
                    nha_may = str(row.get("Nhà máy (*)")).strip() if pd.notna(row.get("Nhà máy (*)")) else ""

                    if not ma:
                        errors.append(f"Dòng {index + 2}: Thiếu mã thiết bị.")
                        continue
                    if not ten:
                        errors.append(f"Dòng {index + 2}: Thiếu tên thiết bị.")
                        continue
                    if not nha_may:
                        errors.append(f"Dòng {index + 2}: Thiếu nhà máy.")
                        continue

                    # Kiểm tra phân quyền nhà máy của người dùng
                    if not has_all_access and user_factory and nha_may != user_factory:
                        nha_may = user_factory

                    # Tìm thiết bị cha nếu có
                    cha_id = None
                    cha_ma_day_du = row.get("Mã đầy đủ thiết bị chi tiết cha")
                    if pd.notna(cha_ma_day_du) and str(cha_ma_day_du).strip():
                        cha_ma_day_du = str(cha_ma_day_du).strip()
                        try:
                            parent = ThietBi.objects.get(ma_day_du=cha_ma_day_du)
                            # Kiểm tra xem cha có thuộc nhà máy được phân quyền không
                            _ensure_thiet_bi_access(request.user, parent)
                            cha_id = parent.id
                        except ThietBi.DoesNotExist:
                            errors.append(f"Dòng {index + 2}: Thiết bị cha '{cha_ma_day_du}' không tồn tại.")
                            continue
                        except PermissionDenied:
                            errors.append(f"Dòng {index + 2}: Bạn không có quyền thao tác với thiết bị cha '{cha_ma_day_du}'.")
                            continue

                    # Tính mã đầy đủ để check trùng hoặc update
                    if cha_id:
                        parent = ThietBi.objects.get(id=cha_id)
                        ma_day_du = f"{parent.ma_day_du}.{ma}"
                    else:
                        ma_day_du = ma

                    # Check quyền tương ứng (create vs update)
                    thiet_bi = ThietBi.objects.filter(ma_day_du=ma_day_du).first()
                    is_new = thiet_bi is None

                    if is_new:
                        if not has_profile_permission(request.user, "can_create_equipment"):
                            errors.append(f"Dòng {index + 2}: Bạn không có quyền tạo mới thiết bị.")
                            continue
                        thiet_bi = ThietBi(ma_day_du=ma_day_du)
                    else:
                        if not has_profile_permission(request.user, "can_edit_equipment"):
                            errors.append(f"Dòng {index + 2}: Bạn không có quyền sửa thiết bị '{ma_day_du}'.")
                            continue
                        _ensure_thiet_bi_access(request.user, thiet_bi)

                    # Điền các giá trị
                    thiet_bi.ma = ma
                    thiet_bi.ten = ten
                    thiet_bi.cha_id = cha_id
                    thiet_bi.nha_may = nha_may

                    thiet_bi.loai = str(row.get("Loại/Phân loại")).strip() if pd.notna(row.get("Loại/Phân loại")) else ""
                    thiet_bi.trang_thai = str(row.get("Trạng thái")).strip() if pd.notna(row.get("Trạng thái")) else ""
                    thiet_bi.nha_che_tao = str(row.get("Nhà chế tạo")).strip() if pd.notna(row.get("Nhà chế tạo")) else ""
                    thiet_bi.nha_cung_cap = str(row.get("Nhà cung cấp")).strip() if pd.notna(row.get("Nhà cung cấp")) else ""
                    thiet_bi.nuoc_san_xuat = str(row.get("Nước sản xuất")).strip() if pd.notna(row.get("Nước sản xuất")) else ""
                    thiet_bi.ma_van_hanh = str(row.get("Mã vận hành")).strip() if pd.notna(row.get("Mã vận hành")) else ""
                    thiet_bi.bo_phan_quan_ly = str(row.get("Bộ phận quản lý")).strip() if pd.notna(row.get("Bộ phận quản lý")) else ""
                    thiet_bi.bang_ve = str(row.get("Bảng vẽ")).strip() if pd.notna(row.get("Bảng vẽ")) else ""
                    thiet_bi.mo_ta_ky_thuat = str(row.get("Thông số kỹ thuật")).strip() if pd.notna(row.get("Thông số kỹ thuật")) else ""

                    # Xử lý ngày
                    def parse_date(val):
                        if pd.isna(val) or not str(val).strip():
                            return None
                        try:
                            if isinstance(val, datetime):
                                return val.date()
                            if isinstance(val, pd.Timestamp):
                                return val.date()
                            return pd.to_datetime(str(val).strip()).date()
                        except Exception:
                            return None

                    thiet_bi.ngay_lap_dat = parse_date(row.get("Ngày lắp đặt (YYYY-MM-DD)"))
                    thiet_bi.ngay_dua_vao_van_hanh = parse_date(row.get("Ngày vận hành (YYYY-MM-DD)"))

                    thiet_bi.save()
                    success_count += 1
                except Exception as e:
                    errors.append(f"Dòng {index + 2}: Lỗi hệ thống: {str(e)}")
                    continue

        if errors:
            return Response({
                'message': f'Import hoàn tất nhưng có một số lỗi. Thành công: {success_count}/{len(df)} dòng.',
                'success_count': success_count,
                'errors': errors
            }, status=status.HTTP_207_MULTI_STATUS)
            
        return Response({
            'message': f'Import thành công {success_count}/{len(df)} thiết bị.',
            'success_count': success_count,
            'errors': []
        }, status=status.HTTP_200_OK)
