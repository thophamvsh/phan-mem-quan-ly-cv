import io
import pandas as pd
from datetime import datetime, time
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from core.factory_scope import (
    filter_queryset_by_factory,
    get_user_factory_name,
    get_user_factory_code,
    has_profile_permission,
    has_all_factory_access,
)
from .models import ThietBi, ThongSoToMay
from .services.thongso_tomay_service import get_specific_thiet_bi


def _excel_template_tomay(request, device_type):
    """Logic cốt lõi tạo template Excel cho thông số tổ máy H1 hoặc H2"""
    try:
        factory_code = request.query_params.get("factory_code")
        if not factory_code or not has_all_factory_access(request.user):
            factory_code = get_user_factory_code(request.user) or "SH"
        is_vs = factory_code.startswith("VS")

        # Tạo 24 chu kỳ từ 00:00 đến 23:00 (mỗi giờ)
        time_cycles = []
        for hour in range(24):
            time_cycles.append(time(hour, 0).strftime("%H:%M"))

        if is_vs:
            # 23 thông số của Vĩnh Sơn
            vs_thong_so = [
                {"ten": "Nhiệt độ dầu", "ma": "nhiet_do_dau_o_do", "don_vi": "°C", "nhom": "Ổ đỡ"},
                {"ten": "Nhiệt độ ổ đỡ", "ma": "nhiet_do_o_do", "don_vi": "°C", "nhom": "Ổ đỡ"},
                {"ten": "Nhiệt độ ổ hướng", "ma": "nhiet_do_o_huong_o_do", "don_vi": "°C", "nhom": "Ổ đỡ"},
                {"ten": "Cuộn dây 1", "ma": "nhiet_do_cuon_day_stato_1", "don_vi": "°C", "nhom": "Stator"},
                {"ten": "Cuộn dây 2", "ma": "nhiet_do_cuon_day_stato_2", "don_vi": "°C", "nhom": "Stator"},
                {"ten": "Lõi sắt 1", "ma": "nhiet_do_loi_sat_stato_1", "don_vi": "°C", "nhom": "Stator"},
                {"ten": "Lõi sắt 2", "ma": "nhiet_do_loi_sat_stato_2", "don_vi": "°C", "nhom": "Stator"},
                {"ten": "Nước đầu vào", "ma": "nuoc_lam_mat_dau_vao", "don_vi": "°C", "nhom": "Nước làm mát"},
                {"ten": "Nước đầu ra", "ma": "nuoc_lam_mat_dau_ra", "don_vi": "°C", "nhom": "Nước làm mát"},
                {"ten": "Gió làm mát", "ma": "nhiet_do_khi_mat", "don_vi": "°C", "nhom": "Nước làm mát"},
                {"ten": "Nhiệt độ dầu", "ma": "nhiet_do_dau_o_huong_tuabin", "don_vi": "°C", "nhom": "Ổ hướng tuabin"},
                {"ten": "Áp suất dầu", "ma": "ap_suat_dau_o_huong_tuabin", "don_vi": "bar", "nhom": "Ổ hướng tuabin"},
                {"ten": "Ổ hướng 1", "ma": "nhiet_do_o_huong_1_tuabin", "don_vi": "°C", "nhom": "Ổ hướng tuabin"},
                {"ten": "Ổ hướng 2", "ma": "nhiet_do_o_huong_2_tuabin", "don_vi": "°C", "nhom": "Ổ hướng tuabin"},
                {"ten": "Áp suất dầu", "ma": "ap_suat_dau_thuy_luc", "don_vi": "bar", "nhom": "Dầu thủy lực"},
                {"ten": "Mức dầu", "ma": "muc_dau_thuy_luc", "don_vi": "DB", "nhom": "Dầu thủy lực"},
                {"ten": "Nhiệt độ dầu", "ma": "nhiet_do_dau_thuy_luc", "don_vi": "°C", "nhom": "Dầu thủy lực"},
                {"ten": "Tốc độ", "ma": "toc_do", "don_vi": "v/ph", "nhom": "Hệ thống điều tốc"},
                {"ten": "Giới hạn độ mở", "ma": "gioi_han_do_mo", "don_vi": "%", "nhom": "Hệ thống điều tốc"},
                {"ten": "Độ mở kim", "ma": "do_mo_kim", "don_vi": "%", "nhom": "Hệ thống điều tốc"},
                {"ten": "Độ mở cánh hướng", "ma": "do_mo_canh_huong", "don_vi": "%", "nhom": "Hệ thống điều tốc"},
                {"ten": "Độ giảm tốc", "ma": "do_giam_toc", "don_vi": "BPO", "nhom": "Hệ thống điều tốc"},
                {"ten": "Độ gia tăng tần số", "ma": "do_gia_tang_tan_so", "don_vi": "CHF0", "nhom": "Hệ thống điều tốc"},
            ]

            # Tạo DataFrame rỗng
            data = []
            row1 = [""] * 24
            row2 = ["Thời gian (giờ)"] + [ts["ten"] for ts in vs_thong_so]
            row3 = [""] + [ts["don_vi"] for ts in vs_thong_so]

            data.append(row1)
            data.append(row2)
            data.append(row3)

            for i in range(24):
                row = [i + 1] + [""] * 23
                data.append(row)

            df = pd.DataFrame(data)
            output = io.BytesIO()
            sheet_name = f"Thông số {device_type}"
            
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

                # A1="NGÀY", C1=Date, E1:X1=Tiêu đề
                worksheet["A1"] = "NGÀY"
                worksheet["A1"].font = Font(name="Times New Roman", bold=True, color="FF0000", size=11)
                worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

                now_str = timezone.localtime(timezone.now()).strftime("%d-%m-%Y")
                worksheet["C1"] = now_str
                worksheet["C1"].font = Font(name="Times New Roman", bold=True, color="FF0000", size=11)
                worksheet["C1"].alignment = Alignment(horizontal="center", vertical="center")

                worksheet.merge_cells("E1:X1")
                worksheet["E1"] = f"BẢNG GHI THÔNG SỐ VẬN HÀNH CƠ CỦA TỔ MÁY {device_type}"
                worksheet["E1"].font = Font(name="Times New Roman", bold=True, color="0000FF", size=14)
                worksheet["E1"].alignment = Alignment(horizontal="center", vertical="center")

                # Merge A2:A3 cho "Thời gian (giờ)"
                worksheet.merge_cells("A2:A3")
                worksheet["A2"] = "Thời\ngian\n(giờ)"
                worksheet["A2"].font = Font(name="Times New Roman", bold=True, size=10)
                worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # Merge các nhóm
                worksheet.merge_cells("B2:D2")
                worksheet["B2"] = "Ổ đỡ"
                worksheet.merge_cells("E2:H2")
                worksheet["E2"] = "Stator"
                worksheet.merge_cells("I2:K2")
                worksheet["I2"] = "Nước làm mát"
                worksheet.merge_cells("L2:O2")
                worksheet["L2"] = "Ổ hướng tuabin"
                worksheet.merge_cells("P2:R2")
                worksheet["P2"] = "Dầu thủy lực"
                worksheet.merge_cells("S2:X2")
                worksheet["S2"] = "Hệ thống điều tốc"

                pastel_fills = {
                    "Ổ đỡ": PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid"),
                    "Stator": PatternFill(start_color="E6F9F2", end_color="E6F9F2", fill_type="solid"),
                    "Nước làm mát": PatternFill(start_color="E6F7F7", end_color="E6F7F7", fill_type="solid"),
                    "Ổ hướng tuabin": PatternFill(start_color="F2EBF9", end_color="F2EBF9", fill_type="solid"),
                    "Dầu thủy lực": PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid"),
                    "Hệ thống điều tốc": PatternFill(start_color="FFFFE6", end_color="FFFFE6", fill_type="solid"),
                }

                # Style Row 2 & Row 3
                for col in range(2, 25):
                    ts_info = vs_thong_so[col - 2]
                    nhom = ts_info["nhom"]
                    group_fill = pastel_fills.get(nhom)

                    cell2 = worksheet.cell(row=2, column=col)
                    cell2.font = Font(name="Times New Roman", bold=True, size=10)
                    cell2.alignment = Alignment(horizontal="center", vertical="center")
                    if group_fill:
                        cell2.fill = group_fill

                    cell3 = worksheet.cell(row=3, column=col)
                    cell3.font = Font(name="Times New Roman", bold=True, size=9)
                    cell3.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if group_fill:
                        cell3.fill = group_fill

                gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                worksheet["A2"].fill = gray_fill
                worksheet["A3"].fill = gray_fill

                # Style B4:X27 (yellow data cells)
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                for row in range(4, 28):
                    cell_a = worksheet.cell(row=row, column=1)
                    cell_a.font = Font(name="Times New Roman", bold=True, size=10)
                    cell_a.alignment = Alignment(horizontal="center", vertical="center")

                    for col in range(2, 25):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = Font(name="Times New Roman", size=10)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.fill = yellow_fill

                # Border & Dimensions
                thin_border = Border(
                    left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"),
                )
                for row in range(1, 28):
                    for col in range(1, 25):
                        worksheet.cell(row=row, column=col).border = thin_border

                worksheet.column_dimensions["A"].width = 10
                for col in range(2, 25):
                    col_letter = chr(64 + col)
                    worksheet.column_dimensions[col_letter].width = 12

                worksheet.row_dimensions[1].height = 28
                worksheet.row_dimensions[2].height = 24
                worksheet.row_dimensions[3].height = 36
                for r in range(4, 28):
                    worksheet.row_dimensions[r].height = 20

            file_content = output.getvalue()
            filename = "ThongSoToMay_Template.xlsx" if device_type == "H1" else "ThongSoToMayH2_Template.xlsx"

            response = HttpResponse(
                file_content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            response["Access-Control-Allow-Origin"] = "*"
            response["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
            response["Access-Control-Allow-Headers"] = "Content-Type"
            return response

        else:
            # Sông Hinh (SH) layout gốc
            thong_so_to_may = [
                {"ten": "Áp lực nước", "ma": "ap_luc_nuoc", "don_vi": "bar"},
                {"ten": "Áp lực chèn trục", "ma": "ap_luc_chen_truc", "don_vi": "bar"},
                {"ten": "Lưu lượng chèn trục", "ma": "luu_luong_chen_truc", "don_vi": "l/p"},
                {"ten": "Lưu lượng ổ hướng tuabin", "ma": "luu_luong_o_huong_tuabin", "don_vi": "l/p"},
                {"ten": "Nhiệt độ ổ hướng tuabin", "ma": "nhiet_do_o_huong_tuabin", "don_vi": "°C"},
                {"ten": "Lưu lượng ổ hướng máy phát", "ma": "luu_luong_o_huong_may_phat", "don_vi": "l/p"},
                {"ten": "Nhiệt độ ổ hướng máy phát", "ma": "nhiet_do_o_huong_may_phat", "don_vi": "°C"},
                {"ten": "Lưu lượng ổ đỡ máy phát", "ma": "luu_luong_o_do_may_phat", "don_vi": "l/p"},
                {"ten": "Nhiệt độ ổ đỡ", "ma": "nhiet_do_o_do", "don_vi": "°C"},
                {"ten": "Nhiệt độ ổ hướng - ổ đỡ", "ma": "nhiet_do_o_huong_o_do", "don_vi": "°C"},
                {"ten": "Nhiệt độ đầu ổ đỡ", "ma": "nhiet_do_dau_o_do", "don_vi": "°C"},
                {"ten": "Lưu lượng làm mát máy phát", "ma": "luu_luong_lam_mat_may_phat", "don_vi": "l/p"},
                {"ten": "Nhiệt độ nước làm mát máy phát", "ma": "nhiet_do_nuoc_lam_mat_may_phat", "don_vi": "°C"},
                {"ten": "Nhiệt độ khí mát", "ma": "nhiet_do_khi_mat", "don_vi": "°C"},
                {"ten": "Nhiệt độ khí nóng", "ma": "nhiet_do_khi_nong", "don_vi": "°C"},
                {"ten": "Nhiệt độ cuộn dây stato", "ma": "nhiet_do_cuon_day_stato", "don_vi": "°C"},
                {"ten": "Tốc độ", "ma": "toc_do", "don_vi": "v/ph"},
                {"ten": "Giới hạn độ mở cánh hướng", "ma": "gioi_han_do_mo_canh_huong", "don_vi": "%"},
                {"ten": "Độ mở cánh hướng", "ma": "do_mo_canh_huong", "don_vi": "%"},
                {"ten": "Độ rơi tốc", "ma": "do_roi_toc", "don_vi": "%"},
            ]

            data = []
            header_row1 = [f"THÔNG SỐ {device_type}"] + [""] * (len(thong_so_to_may) - 1)
            header_row2 = [ts["ten"] for ts in thong_so_to_may]
            header_row3 = [ts["don_vi"] for ts in thong_so_to_may]

            data.append(header_row1)
            data.append(header_row2)
            data.append(header_row3)

            for i in range(24):
                row = ["-"] * len(thong_so_to_may)
                data.append(row)

            df = pd.DataFrame(data)
            output = io.BytesIO()
            sheet_name = f"Thông số {device_type}"
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
                worksheet.merge_cells("A1:T1")

                header_font = Font(bold=True, size=14)
                header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")

                worksheet["A1"].font = header_font
                worksheet["A1"].fill = header_fill
                worksheet["A1"].alignment = header_alignment

                for col in range(1, len(thong_so_to_may) + 1):
                    cell = worksheet.cell(row=2, column=col)
                    cell.font = Font(bold=True, size=12)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

                for col in range(1, len(thong_so_to_may) + 1):
                    cell = worksheet.cell(row=3, column=col)
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

                for row in range(4, 28):
                    cell = worksheet.cell(row=row, column=1)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                worksheet.column_dimensions["A"].width = 12
                for col in range(1, len(thong_so_to_may) + 1):
                    worksheet.column_dimensions[chr(64 + col)].width = 15

                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                for row in range(1, 28):
                    for col in range(1, len(thong_so_to_may) + 1):
                        worksheet.cell(row=row, column=col).border = thin_border

            file_content = output.getvalue()
            filename = "ThongSoToMay_Template.xlsx" if device_type == "H1" else "ThongSoToMayH2_Template.xlsx"

            response = HttpResponse(
                file_content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            response["Access-Control-Allow-Origin"] = "*"
            response["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
            response["Access-Control-Allow-Headers"] = "Content-Type"
            return response

    except Exception as e:
        return HttpResponse(f"Lỗi khi tạo template: {str(e)}", status=500)


def _import_excel_tomay(request, device_type=None):
    """Logic cốt lõi import dữ liệu từ Excel cho thông số tổ máy H1/H2"""
    try:
        excel_file = request.FILES.get("file")
        if not excel_file:
            return HttpResponse("Không có file Excel", status=400)

        request_data = getattr(request, "data", request.POST)
        selected_date = request_data.get("selected_date")
        if not selected_date:
            return HttpResponse("Không có ngày được chọn", status=400)

        try:
            target_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
        except ValueError:
            return HttpResponse("Định dạng ngày không hợp lệ", status=400)

        df = pd.read_excel(excel_file, header=None)

        if len(df) < 3:
            return JsonResponse(
                {
                    "error": (
                        f"File Excel không hợp lệ: File phải có ít nhất 3 hàng header. "
                        f"File hiện tại có {len(df)} hàng."
                    ),
                    "status": "error",
                    "expected_rows": 3,
                    "actual_rows": len(df),
                },
                status=400,
            )

        # Tự động nhận diện cấu trúc Vĩnh Sơn (24 cột) hoặc Sông Hinh (20 cột)
        device_code = request_data.get("device_code")
        target_factory = None
        if device_code:
            if device_code.startswith("VS"):
                target_factory = "VS"
            elif device_code.startswith("SH"):
                target_factory = "SH"
                
        if not target_factory:
            req_factory = request_data.get("factory_code")
            if req_factory:
                target_factory = req_factory
                
        if not target_factory:
            target_factory = get_user_factory_code(request.user)

        if target_factory:
            is_vs = (target_factory == "VS")
        else:
            is_vs = False
            if len(df.columns) >= 24:
                first_rows_text = ""
                for r_idx in range(min(3, len(df))):
                    row_vals = [str(x) for x in df.iloc[r_idx].tolist() if pd.notna(x)]
                    first_rows_text += " " + " ".join(row_vals)
                first_rows_text = first_rows_text.upper()
                
                if "SÔNG HINH" in first_rows_text or "SONG HINH" in first_rows_text:
                    is_vs = False
                else:
                    is_vs = True

        expected_columns = 24 if is_vs else 20
        if len(df.columns) < expected_columns:
            return JsonResponse(
                {
                    "error": (
                        f"File Excel không hợp lệ: File phải có ít nhất {expected_columns} cột. "
                        f"File hiện tại có {len(df.columns)} cột."
                    ),
                    "status": "error",
                    "expected_columns": expected_columns,
                    "actual_columns": len(df.columns),
                },
                status=400,
            )

        # Nhận diện tổ máy (H1/H2) từ toàn bộ dòng 1 để tránh lệch ô/cột
        row1_vals = [str(x) for x in df.iloc[0].tolist() if pd.notna(x)]
        header_row1_upper = " ".join(row1_vals).upper()

        if "VẬN HÀNH" in header_row1_upper and not is_vs:
            return JsonResponse(
                {
                    "error": (
                        "File Excel không hợp lệ: Đây là file Thông số Vận hành, không phải file Thông số Tổ máy. "
                        "Vui lòng sử dụng file template Thông số Tổ máy."
                    ),
                    "status": "error",
                    "detected_type": "thong_so_van_hanh",
                },
                status=400,
            )

        detected_type = None
        if "H1" in header_row1_upper:
            detected_type = "H1"
        elif "H2" in header_row1_upper:
            detected_type = "H2"

        if not detected_type:
            line1_text = ", ".join([f"{chr(65+i)}: {str(val)}" for i, val in enumerate(df.iloc[0].tolist()) if pd.notna(val)])
            return JsonResponse(
                {
                    "error": (
                        f'File Excel không hợp lệ: Tiêu đề dòng 1 phải chứa "H1" hoặc "H2" để xác định tổ máy. '
                        f'Các giá trị đọc được ở dòng 1: [{line1_text}].'
                    ),
                    "status": "error",
                },
                status=400,
            )

        if device_type and detected_type != device_type:
            return JsonResponse(
                {
                    "error": (
                        f"File Excel tải lên ({detected_type}) không khớp với "
                        f"tổ máy được chọn trên giao diện ({device_type})."
                    ),
                    "status": "error",
                },
                status=400,
            )

        factory_code = None
        if device_code:
            if device_code.startswith("VS"):
                factory_code = "VS"
            elif device_code.startswith("SH"):
                factory_code = "SH"
        if not factory_code:
            factory_code = request_data.get("factory_code")
            
        user_factory = get_user_factory_code(request.user)
        if user_factory and not has_all_factory_access(request.user):
            factory_code = user_factory
            
        if not factory_code:
            factory_code = "VS" if is_vs else "SH"
        device_code = request_data.get("device_code")

        if device_code:
            if ".H1." in device_code:
                expected_type = "H1"
            elif ".H2." in device_code:
                expected_type = "H2"
            else:
                expected_type = None

            if expected_type and detected_type != expected_type:
                return JsonResponse(
                    {
                        "error": (
                            f"File Excel tải lên ({detected_type}) không khớp với "
                            f"tổ máy được chọn trên giao diện ({expected_type})."
                        ),
                        "status": "error",
                    },
                    status=400,
                )
        else:
            device_code = f"{factory_code}.TB.{detected_type}.GE"

        # Định nghĩa các thông số mong đợi
        if is_vs:
            expected_params = [
                "Nhiệt độ dầu", "Nhiệt độ ổ đỡ", "Nhiệt độ ổ hướng",
                "Cuộn dây 1", "Cuộn dây 2", "Lõi sắt 1", "Lõi sắt 2",
                "Nước đầu vào", "Nước đầu ra", "Gió làm mát",
                "Nhiệt độ dầu", "Áp suất dầu", "Ổ hướng 1", "Ổ hướng 2",
                "Áp suất dầu", "Mức dầu", "Nhiệt độ dầu",
                "Tốc độ", "Giới hạn độ mở", "Độ mở kim", "Độ mở cánh hướng", "Độ giảm tốc", "Độ gia tăng tần số"
            ]
            header_check_row = df.iloc[2, 1:24].fillna("").astype(str).tolist()
        else:
            expected_params = [
                "Áp lực nước", "Áp lực chèn trục", "Lưu lượng chèn trục", "Lưu lượng ổ hướng tuabin",
                "Nhiệt độ ổ hướng tuabin", "Lưu lượng ổ hướng máy phát", "Nhiệt độ ổ hướng máy phát",
                "Lưu lượng ổ đỡ máy phát", "Nhiệt độ ổ đỡ", "Nhiệt độ ổ hướng - ổ đỡ",
                "Nhiệt độ đầu ổ đỡ", "Lưu lượng làm mát máy phát", "Nhiệt độ nước làm mát máy phát",
                "Nhiệt độ khí mát", "Nhiệt độ khí nóng", "Nhiệt độ cuộn dây stato",
                "Tốc độ", "Giới hạn độ mở cánh hướng", "Độ mở cánh hướng", "Độ rơi tốc"
            ]
            header_check_row = df.iloc[1, :20].fillna("").astype(str).tolist()

        # Kiểm tra tính chính xác của các cột
        missing_params = []
        for idx, expected_param in enumerate(expected_params):
            if idx >= len(header_check_row):
                missing_params.append(f'Cột {chr(65 + (idx + 1 if is_vs else idx))}: Thiếu "{expected_param}"')
                continue

            actual_value = str(header_check_row[idx]).strip()
            if not actual_value:
                missing_params.append(f'Cột {chr(65 + (idx + 1 if is_vs else idx))}: Trống, mong đợi "{expected_param}"')
                continue

            expected_normalized = " ".join(expected_param.lower().strip().split())
            actual_normalized = " ".join(actual_value.lower().strip().split())

            if expected_normalized == actual_normalized or expected_normalized in actual_normalized or actual_normalized in expected_normalized:
                continue

            # Fuzzy matching
            expected_words = expected_normalized.split()
            actual_words = actual_normalized.split()

            def simple_edit_distance(word1, word2):
                if abs(len(word1) - len(word2)) > 2:
                    return float('inf')
                if word1 == word2:
                    return 0
                min_len = min(len(word1), len(word2))
                max_len = max(len(word1), len(word2))
                diff = 0
                for i in range(min_len):
                    if word1[i] != word2[i]:
                        diff += 1
                diff += (max_len - min_len)
                return diff

            words_match = 0
            total_words = len(expected_words)
            if total_words > 0:
                for exp_word in expected_words:
                    best_match = min(
                        [simple_edit_distance(exp_word, act_word) for act_word in actual_words],
                        default=float('inf')
                    )
                    if best_match <= 2:
                        words_match += 1
                if words_match / total_words >= 0.8:
                    continue

            missing_params.append(f'Cột {chr(65 + (idx + 1 if is_vs else idx))}: Mong đợi "{expected_param}" nhưng có "{actual_value}"')

        if missing_params:
            return JsonResponse(
                {
                    "error": (
                        f"File Excel không hợp lệ: Các cột không khớp với template {detected_type}.\n"
                        + "\n".join(missing_params[:5])
                        + (f"\n... và {len(missing_params) - 5} cột khác" if len(missing_params) > 5 else "")
                    ),
                    "status": "error",
                    "missing_params": missing_params[:10],
                },
                status=400,
            )

        # Cấu trúc ánh xạ cột sang model field
        if is_vs:
            column_mapping = {
                1: {"ten": "Nhiệt độ dầu", "ma": "nhiet_do_dau_o_do", "don_vi": "°C"},
                2: {"ten": "Nhiệt độ ổ đỡ", "ma": "nhiet_do_o_do", "don_vi": "°C"},
                3: {"ten": "Nhiệt độ ổ hướng", "ma": "nhiet_do_o_huong_o_do", "don_vi": "°C"},
                4: {"ten": "Cuộn dây 1", "ma": "nhiet_do_cuon_day_stato_1", "don_vi": "°C"},
                5: {"ten": "Cuộn dây 2", "ma": "nhiet_do_cuon_day_stato_2", "don_vi": "°C"},
                6: {"ten": "Lõi sắt 1", "ma": "nhiet_do_loi_sat_stato_1", "don_vi": "°C"},
                7: {"ten": "Lõi sắt 2", "ma": "nhiet_do_loi_sat_stato_2", "don_vi": "°C"},
                8: {"ten": "Nước đầu vào", "ma": "nuoc_lam_mat_dau_vao", "don_vi": "°C"},
                9: {"ten": "Nước đầu ra", "ma": "nuoc_lam_mat_dau_ra", "don_vi": "°C"},
                10: {"ten": "Gió làm mát", "ma": "nhiet_do_khi_mat", "don_vi": "°C"},
                11: {"ten": "Nhiệt độ dầu", "ma": "nhiet_do_dau_o_huong_tuabin", "don_vi": "°C"},
                12: {"ten": "Áp suất dầu", "ma": "ap_suat_dau_o_huong_tuabin", "don_vi": "bar"},
                13: {"ten": "Ổ hướng 1", "ma": "nhiet_do_o_huong_1_tuabin", "don_vi": "°C"},
                14: {"ten": "Ổ hướng 2", "ma": "nhiet_do_o_huong_2_tuabin", "don_vi": "°C"},
                15: {"ten": "Áp suất dầu", "ma": "ap_suat_dau_thuy_luc", "don_vi": "bar"},
                16: {"ten": "Mức dầu", "ma": "muc_dau_thuy_luc", "don_vi": "DB"},
                17: {"ten": "Nhiệt độ dầu", "ma": "nhiet_do_dau_thuy_luc", "don_vi": "°C"},
                18: {"ten": "Tốc độ", "ma": "toc_do", "don_vi": "v/ph"},
                19: {"ten": "Giới hạn độ mở", "ma": "gioi_han_do_mo", "don_vi": "%"},
                20: {"ten": "Độ mở kim", "ma": "do_mo_kim", "don_vi": "%"},
                21: {"ten": "Độ mở cánh hướng", "ma": "do_mo_canh_huong", "don_vi": "%"},
                22: {"ten": "Độ giảm tốc", "ma": "do_giam_toc", "don_vi": "BPO"},
                23: {"ten": "Độ gia tăng tần số", "ma": "do_gia_tang_tan_so", "don_vi": "CHF0"},
            }
        else:
            column_mapping = {
                i: {"ten": p, "ma": m, "don_vi": d}
                for i, (p, m, d) in enumerate(
                    [
                        ("Áp lực nước", "ap_luc_nuoc", "bar"),
                        ("Áp lực chèn trục", "ap_luc_chen_truc", "bar"),
                        ("Lưu lượng chèn trục", "luu_luong_chen_truc", "l/p"),
                        ("Lưu lượng ổ hướng tuabin", "luu_luong_o_huong_tuabin", "l/p"),
                        ("Nhiệt độ ổ hướng tuabin", "nhiet_do_o_huong_tuabin", "°C"),
                        ("Lưu lượng ổ hướng máy phát", "luu_luong_o_huong_may_phat", "l/p"),
                        ("Nhiệt độ ổ hướng máy phát", "nhiet_do_o_huong_may_phat", "°C"),
                        ("Lưu lượng ổ đỡ máy phát", "luu_luong_o_do_may_phat", "l/p"),
                        ("Nhiệt độ ổ đỡ", "nhiet_do_o_do", "°C"),
                        ("Nhiệt độ ổ hướng - ổ đỡ", "nhiet_do_o_huong_o_do", "°C"),
                        ("Nhiệt độ đầu ổ đỡ", "nhiet_do_dau_o_do", "°C"),
                        ("Lưu lượng làm mát máy phát", "luu_luong_lam_mat_may_phat", "l/p"),
                        ("Nhiệt độ nước làm mát máy phát", "nhiet_do_nuoc_lam_mat_may_phat", "°C"),
                        ("Nhiệt độ khí mát", "nhiet_do_khi_mat", "°C"),
                        ("Nhiệt độ khí nóng", "nhiet_do_khi_nong", "°C"),
                        ("Nhiệt độ cuộn dây stato", "nhiet_do_cuon_day_stato", "°C"),
                        ("Tốc độ", "toc_do", "v/ph"),
                        ("Giới hạn độ mở cánh hướng", "gioi_han_do_mo_canh_huong", "%"),
                        ("Độ mở cánh hướng", "do_mo_canh_huong", "%"),
                        ("Độ rơi tốc", "do_roi_toc", "%"),
                    ]
                )
            }

        try:
            thiet_bi = filter_queryset_by_factory(
                ThietBi.objects.all(),
                request.user,
                "nha_may",
                "string",
            ).get(ma_day_du=device_code)
        except ThietBi.DoesNotExist:
            return HttpResponse(f"Không tìm thấy thiết bị gốc {device_code} trong quyền hạn", status=400)

        import pytz
        vietnam_tz = pytz.timezone("Asia/Ho_Chi_Minh")
        time_cycles = []
        for hour in range(24):
            naive_dt = datetime.combine(target_date, time(hour, 0))
            time_cycles.append(vietnam_tz.localize(naive_dt))

        imported_count = 0
        prefix = ".".join(thiet_bi.ma_day_du.split(".")[:3])  # VS.TB.H1 / SH.TB.H1
        existing_records = ThongSoToMay.objects.filter(
            thiet_bi__ma_day_du__startswith=prefix, ngay_nhap=target_date
        )

        if not request.user.is_superuser:
            for rec in existing_records:
                if rec.nguoi_nhap and rec.nguoi_nhap != request.user:
                    return JsonResponse({
                        'error': f'Bạn không có quyền cập nhật thông số ngày {target_date} vì một số bản ghi đã được nhập bởi người dùng khác ({rec.nguoi_nhap.username or rec.nguoi_nhap.email}).',
                        'status': 'error'
                    }, status=403)

        existing_lookup = {
            (rec.ma_thong_so or rec.ten_thong_so, rec.thoi_diem_nhap): rec for rec in existing_records
        }

        to_create = []
        to_update = []

        start_col = 1 if is_vs else 0
        end_col = 24 if is_vs else 20

        # Đọc dữ liệu từ hàng 4 (chỉ số 3) đến hàng 27 (chỉ số 26)
        for row_idx in range(3, 27):
            time_cycle = time_cycles[row_idx - 3] if row_idx - 3 < len(time_cycles) else None
            if not time_cycle:
                continue

            for col_idx in range(start_col, end_col):
                if col_idx not in column_mapping:
                    continue

                mapping = column_mapping[col_idx]
                value = df.iloc[row_idx, col_idx] if row_idx < len(df) else None

                # Cho phép lưu trữ chữ và số kết hợp
                if pd.isna(value) or value == "" or value == "-":
                    store_value = None
                else:
                    val_str = str(value).strip()
                    try:
                        # Thử parse sang float để chuẩn hóa các số
                        float_val = float(val_str.replace(",", "."))
                        if float_val.is_integer():
                            store_value = str(int(float_val))
                        else:
                            store_value = str(float_val)
                    except (ValueError, TypeError):
                        store_value = val_str

                try:
                    nha_may_val = get_user_factory_name(request.user) or thiet_bi.nha_may
                    lookup_key = (mapping["ma"], time_cycle)

                    target_tb = get_specific_thiet_bi(thiet_bi, mapping["ma"])

                    if lookup_key in existing_lookup:
                        obj = existing_lookup[lookup_key]
                        if (obj.gia_tri != store_value or 
                            obj.nha_may != nha_may_val or 
                            obj.thiet_bi_id != target_tb.id or 
                            obj.ten_thong_so != mapping["ten"] or 
                            obj.ma_thong_so != mapping["ma"] or
                            obj.nguoi_nhap != request.user):
                            obj.gia_tri = store_value
                            obj.nha_may = nha_may_val
                            obj.thiet_bi = target_tb
                            obj.ten_thong_so = mapping["ten"]
                            obj.ma_thong_so = mapping["ma"]
                            obj.nguoi_nhap = request.user
                            to_update.append(obj)
                    else:
                        obj = ThongSoToMay(
                            thiet_bi=target_tb,
                            ten_thong_so=mapping["ten"],
                            thoi_diem_nhap=time_cycle,
                            ngay_nhap=target_date,
                            ma_thong_so=mapping["ma"],
                            don_vi=mapping["don_vi"],
                            gia_tri=store_value,
                            nha_may=nha_may_val,
                            ky_hieu_van_hanh=f'{detected_type}_{mapping["ma"]}',
                            ghi_chu=f"Import từ Excel - {target_date}",
                            nguoi_nhap=request.user
                        )
                        to_create.append(obj)
                        existing_lookup[lookup_key] = obj

                    imported_count += 1
                except Exception:
                    continue

        if to_create:
            ThongSoToMay.objects.bulk_create(to_create)
        if to_update:
            ThongSoToMay.objects.bulk_update(to_update, ["gia_tri", "nha_may", "thiet_bi", "ten_thong_so", "ma_thong_so", "nguoi_nhap"])

        return JsonResponse(
            {
                "message": f"Import thành công {imported_count} bản ghi thông số tổ máy {detected_type}",
                "imported_count": imported_count,
                "status": "success",
            },
            status=200,
        )

    except Exception as e:
        return JsonResponse({"error": f"Lỗi khi import: {str(e)}", "status": "error"}, status=500)


# ==========================================
# PUBLIC API ENDPOINTS (DECORATED WITH DRF)
# ==========================================

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def excel_template_tomay(request):
    """Tạo template Excel chung cho H1/H2, nhận diện qua param ?type=H1|H2"""
    if not has_profile_permission(request.user, "can_export_excel"):
        return JsonResponse(
            {
                "error": (
                    "Tài khoản của bạn chưa được cấp quyền xuất dữ liệu "
                    "hoặc tải template Excel. Vui lòng liên hệ quản trị viên."
                )
            },
            status=403,
        )
    device_type = request.query_params.get("type", "H1").upper()
    if device_type not in ["H1", "H2"]:
        return JsonResponse(
            {"error": "Tổ máy không hợp lệ (chỉ hỗ trợ H1 hoặc H2)"},
            status=400,
        )
    return _excel_template_tomay(request, device_type)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def excel_template_h1(request):
    """Wrapper tương thích ngược cho template H1"""
    if not has_profile_permission(request.user, "can_export_excel"):
        return JsonResponse(
            {
                "error": (
                    "Tài khoản của bạn chưa được cấp quyền xuất dữ liệu "
                    "hoặc tải template Excel. Vui lòng liên hệ quản trị viên."
                )
            },
            status=403,
        )
    return _excel_template_tomay(request, "H1")


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def excel_template_h2(request):
    """Wrapper tương thích ngược cho template H2"""
    if not has_profile_permission(request.user, "can_export_excel"):
        return JsonResponse(
            {
                "error": (
                    "Tài khoản của bạn chưa được cấp quyền xuất dữ liệu "
                    "hoặc tải template Excel. Vui lòng liên hệ quản trị viên."
                )
            },
            status=403,
        )
    return _excel_template_tomay(request, "H2")


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_excel_tomay(request):
    """Import dữ liệu từ Excel cho thông số tổ máy (Tự động nhận diện H1/H2)"""
    if not has_profile_permission(request.user, "can_import_excel"):
        return JsonResponse(
            {
                "error": (
                    "Tài khoản của bạn chưa được cấp quyền import dữ liệu từ Excel. "
                    "Vui lòng liên hệ quản trị viên."
                )
            },
            status=403,
        )
    return _import_excel_tomay(request)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_excel_h1(request):
    """Wrapper tương thích ngược cho import H1"""
    if not has_profile_permission(request.user, "can_import_excel"):
        return JsonResponse(
            {
                "error": (
                    "Tài khoản của bạn chưa được cấp quyền import dữ liệu từ Excel. "
                    "Vui lòng liên hệ quản trị viên."
                )
            },
            status=403,
        )
    return _import_excel_tomay(request, device_type="H1")


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_excel_h2(request):
    """Wrapper tương thích ngược cho import H2"""
    if not has_profile_permission(request.user, "can_import_excel"):
        return JsonResponse(
            {
                "error": (
                    "Tài khoản của bạn chưa được cấp quyền import dữ liệu từ Excel. "
                    "Vui lòng liên hệ quản trị viên."
                )
            },
            status=403,
        )
    return _import_excel_tomay(request, device_type="H2")
