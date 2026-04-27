from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

@csrf_exempt
@require_http_methods(["GET"])
def test_endpoint(request):
    """Test endpoint không cần authentication"""
    return JsonResponse({'message': 'Test endpoint working!'})

@csrf_exempt
@require_http_methods(["GET"])
def test_excel(request):
    """Test Excel endpoint"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        import io

        # Tạo workbook mới
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"

        # Tạo header
        ws.cell(row=1, column=1, value="Test Header")
        ws.cell(row=2, column=1, value="Test Data")

        # Tạo buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="test.xlsx"'
        response['Content-Length'] = len(buffer.getvalue())

        return response

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
