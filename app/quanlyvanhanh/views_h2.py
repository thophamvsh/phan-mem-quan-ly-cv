import io
import pandas as pd
from datetime import datetime, time
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from core.factory_scope import filter_queryset_by_factory, get_user_factory_name
from .models import ThietBi, ThongSoToMay


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def excel_template_h2(request):
    """Tạo template Excel cho thông số tổ máy H2"""
    try:
        # Tạo 24 chu kỳ từ 00:00 đến 23:00 (mỗi giờ)
        time_cycles = []
        for hour in range(24):
            time_cycles.append(time(hour, 0).strftime('%H:%M'))

        # Định nghĩa các thông số tổ máy theo hình
        thong_so_to_may = [
            {"ten": "Áp lực nước", "ma": "ap_luc_nuoc", "don_vi": "bar"},
            {"ten": "Áp lực chèn trục", "ma": "ap_luc_chen_truc", "don_vi": "bar"},
            {"ten": "Lưu lượng chèn trục", "ma": "luu_luong_chen_truc", "don_vi": "l/p"},
            {"ten": "Lưu lượng ổ hướng tuabin", "ma": "luu_luong_o_huong_tuabin", "don_vi": "l/p"},
            {"ten": "Nhiệt độ ổ hướng tuabin", "ma": "nhiet_do_o_huong_tuabin", "don_vi": "°C"},
            {"ten": "Lưu lượng ổ hướng máy phát", "ma": "luu_luong_o_huong_may_phat", "don_vi": "l/p"},
            {"ten": "Nhiệt độ ổ hướng máy phát", "ma": "nhiet_do_o_huong_may_phat", "don_vi": "°C"},
            {"ten": "Lưu lượng ổ đỡ máy phát", "ma": "luu_luong_o_do_may_phat", "don_vi": "l/p"},
            {"ten": "Nhiệt độ ổ đỡ", "ma": "nhiet_do_o_do", "don_vi": "°C"},
            {"ten": "Nhiệt độ ổ hướng - ổ đỡ", "ma": "nhiet_do_o_huong_o_do", "don_vi": "°C"},
            {"ten": "Nhiệt độ đầu ổ đỡ", "ma": "nhiet_do_dau_o_do", "don_vi": "°C"},
            {"ten": "Lưu lượng làm mát máy phát", "ma": "luu_luong_lam_mat_may_phat", "don_vi": "l/p"},
            {"ten": "Nhiệt độ nước làm mát máy phát", "ma": "nhiet_do_nuoc_lam_mat_may_phat", "don_vi": "°C"},
            {"ten": "Nhiệt độ khí mát", "ma": "nhiet_do_khi_mat", "don_vi": "°C"},
            {"ten": "Nhiệt độ khí nóng", "ma": "nhiet_do_khi_nong", "don_vi": "°C"},
            {"ten": "Nhiệt độ cuộn dây stato", "ma": "nhiet_do_cuon_day_stato", "don_vi": "°C"},
            {"ten": "Tốc độ", "ma": "toc_do", "don_vi": "v/ph"},
            {"ten": "Giới hạn độ mở cánh hướng", "ma": "gioi_han_do_mo_canh_huong", "don_vi": "%"},
            {"ten": "Độ mở cánh hướng", "ma": "do_mo_canh_huong", "don_vi": "%"},
            {"ten": "Độ rơi tốc", "ma": "do_roi_toc", "don_vi": "%"},
        ]

        # Tạo DataFrame
        data = []

        # Header rows
        header_row1 = ["THÔNG SỐ H2"] + [""] * (len(thong_so_to_may) - 1)  # A1:T1
        header_row2 = [ts["ten"] for ts in thong_so_to_may]  # A2:T2 - 20 tiêu đề
        header_row3 = [ts["don_vi"] for ts in thong_so_to_may]  # A3:T3 - 20 đơn vị

        data.append(header_row1)
        data.append(header_row2)
        data.append(header_row3)

        # Data rows (24 chu kỳ) - bắt đầu từ A4 đến T27

        for i, time_cycle in enumerate(time_cycles):
            row = ["-"] * len(thong_so_to_may)  # A4:T27 - 20 cột dữ liệu, hiển thị "-" cho ô trống
            data.append(row)

        # Tạo DataFrame
        df = pd.DataFrame(data)

        # Tạo Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Thông số H2', index=False, header=False)

            # Lấy workbook và worksheet để format
            workbook = writer.book
            worksheet = writer.sheets['Thông số H2']

            # Merge cells cho header
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

            # Merge A1:T1 cho "THÔNG SỐ H2" (20 cột từ A đến T)
            worksheet.merge_cells('A1:T1')

            # Style cho header
            header_font = Font(bold=True, size=14)
            header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Apply style to A1
            worksheet['A1'].font = header_font
            worksheet['A1'].fill = header_fill
            worksheet['A1'].alignment = header_alignment

            # Style cho row 2 (tên thông số) - cột A đến T
            for col in range(1, len(thong_so_to_may) + 1):  # A to T
                cell = worksheet.cell(row=2, column=col)
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green

            # Style cho row 3 (đơn vị) - cột A đến T
            for col in range(1, len(thong_so_to_may) + 1):  # A to T
                cell = worksheet.cell(row=3, column=col)
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # Light grey

            # Style cho time column (A) - để trống
            for row in range(4, 28):  # A4 to A27
                cell = worksheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # Không fill màu để trống

            # Set column widths
            worksheet.column_dimensions['A'].width = 12
            for col in range(1, len(thong_so_to_may) + 1):  # A to T
                worksheet.column_dimensions[chr(64 + col)].width = 15

            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in range(1, 28):
                for col in range(1, len(thong_so_to_may) + 1):  # A to T
                    worksheet.cell(row=row, column=col).border = thin_border

        file_content = output.getvalue()

        response = HttpResponse(
            file_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="ThongSoToMayH2_Template.xlsx"'
        response['Access-Control-Allow-Origin'] = '*'
        response['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type'

        return response

    except Exception as e:
        return HttpResponse(
            f'Lỗi khi tạo template: {str(e)}',
            status=500
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_excel_h2(request):
    """Import dữ liệu từ Excel cho thông số tổ máy H2"""
    try:
        # Lấy file Excel
        excel_file = request.FILES.get('file')
        if not excel_file:
            return HttpResponse('Không có file Excel', status=400)

        # Lấy thông tin từ form
        request_data = getattr(request, "data", request.POST)
        selected_date = request_data.get('selected_date')
        if not selected_date:
            return HttpResponse('Không có ngày được chọn', status=400)

        # Parse ngày
        try:
            target_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            return HttpResponse('Định dạng ngày không hợp lệ', status=400)

        # Đọc Excel file
        df = pd.read_excel(excel_file, header=None)

        # ===== VALIDATION: Kiểm tra file Excel có phải là file H2 không =====
        # 1. Kiểm tra số hàng và số cột cơ bản
        if len(df) < 27:
            return JsonResponse({
                'error': f'File Excel không hợp lệ: File thông số H2 phải có ít nhất 27 hàng (3 header + 24 data rows). File hiện tại có {len(df)} hàng.',
                'status': 'error',
                'expected_rows': 27,
                'actual_rows': len(df)
            }, status=400)

        if len(df.columns) < 20:
            return JsonResponse({
                'error': f'File Excel không hợp lệ: File thông số H2 phải có ít nhất 20 cột (A đến T). File hiện tại có {len(df.columns)} cột.',
                'status': 'error',
                'expected_columns': 20,
                'actual_columns': len(df.columns)
            }, status=400)

        # 2. Kiểm tra header row 1 (A1) phải chứa "THÔNG SỐ H2"
        header_row1 = df.iloc[0, 0] if len(df) > 0 and len(df.columns) > 0 else None
        if pd.isna(header_row1) or not isinstance(header_row1, str):
            header_row1 = ""

        # Kiểm tra không phải là file thông số vận hành
        if "VẬN HÀNH" in str(header_row1).upper():
            return JsonResponse({
                'error': 'File Excel không hợp lệ: Đây là file Thông số Vận hành, không phải file Thông số H2. Vui lòng sử dụng file template Thông số H2.',
                'status': 'error',
                'detected_type': 'thong_so_van_hanh',
                'expected_type': 'thong_so_h2'
            }, status=400)

        # Kiểm tra header có chứa "THÔNG SỐ H2" hoặc "H2"
        if "THÔNG SỐ H2" not in str(header_row1).upper() and "H2" not in str(header_row1).upper():
            return JsonResponse({
                'error': f'File Excel không hợp lệ: Header row 1 (A1) phải chứa "THÔNG SỐ H2". Giá trị hiện tại: "{header_row1}". Vui lòng sử dụng file template Thông số H2.',
                'status': 'error',
                'expected_header': 'THÔNG SỐ H2',
                'actual_header': str(header_row1)
            }, status=400)

        # 3. Kiểm tra header row 2 (A2:T2) phải chứa các tên thông số H2 đúng
        # Định nghĩa danh sách tên thông số H2 bắt buộc (20 thông số)
        expected_params_h2 = [
            "Áp lực nước",
            "Áp lực chèn trục",
            "Lưu lượng chèn trục",
            "Lưu lượng ổ hướng tuabin",
            "Nhiệt độ ổ hướng tuabin",
            "Lưu lượng ổ hướng máy phát",
            "Nhiệt độ ổ hướng máy phát",
            "Lưu lượng ổ đỡ máy phát",
            "Nhiệt độ ổ đỡ",
            "Nhiệt độ ổ hướng - ổ đỡ",
            "Nhiệt độ đầu ổ đỡ",
            "Lưu lượng làm mát máy phát",
            "Nhiệt độ nước làm mát máy phát",
            "Nhiệt độ khí mát",
            "Nhiệt độ khí nóng",
            "Nhiệt độ cuộn dây stato",
            "Tốc độ",
            "Giới hạn độ mở cánh hướng",
            "Độ mở cánh hướng",
            "Độ rơi tốc"
        ]

        # Kiểm tra header row 2 (index 1) có chứa đúng 20 tên thông số
        if len(df) < 2:
            return JsonResponse({
                'error': 'File Excel không hợp lệ: Thiếu header row 2 (tên thông số).',
                'status': 'error'
            }, status=400)

        header_row2 = df.iloc[1, :20].fillna("").astype(str).tolist() if len(df.columns) >= 20 else []

        # Kiểm tra từng cột trong header row 2
        # Normalize tất cả về lowercase để so sánh không phân biệt hoa thường
        missing_params = []
        for idx, expected_param in enumerate(expected_params_h2):
            if idx >= len(header_row2):
                missing_params.append(f'Cột {chr(65 + idx)}: Thiếu "{expected_param}"')
                continue

            actual_value = str(header_row2[idx]).strip()

            # Nếu ô trống
            if not actual_value or actual_value == "":
                missing_params.append(f'Cột {chr(65 + idx)}: Trống, mong đợi "{expected_param}"')
                continue

            # Normalize cả hai về lowercase hoàn toàn (không phân biệt H/h, N/n, ...)
            # Loại bỏ khoảng trắng thừa và chuẩn hóa
            expected_normalized = " ".join(expected_param.lower().strip().split())
            actual_normalized = " ".join(actual_value.lower().strip().split())

            # So sánh chính xác sau khi normalize
            if expected_normalized == actual_normalized:
                continue  # Khớp hoàn toàn, bỏ qua

            # So sánh substring (cho phép có thêm/thiếu ký tự)
            if expected_normalized in actual_normalized or actual_normalized in expected_normalized:
                continue  # Khớp substring, bỏ qua

            # Fuzzy matching: Cho phép lỗi đánh máy nhỏ (ví dụ: "đầu" vs "dầu")
            # So sánh theo từ khóa chính (bỏ qua stop words)
            expected_words = expected_normalized.split()
            actual_words = actual_normalized.split()

            # Tính số từ khớp (bỏ qua thứ tự)
            expected_set = set(expected_words)
            actual_set = set(actual_words)

            # Tính similarity dựa trên số từ khớp
            if len(expected_set) > 0:
                similarity = len(expected_set & actual_set) / len(expected_set)
                # Cho phép nếu >= 80% từ khớp (cho phép 1-2 từ khác biệt)
                if similarity >= 0.8:
                    continue

            # Tính edit distance đơn giản cho các từ quan trọng
            # Cho phép sai lệch 1-2 ký tự trong mỗi từ
            def simple_edit_distance(word1, word2):
                """Tính edit distance đơn giản (Levenshtein-like) cho fuzzy matching"""
                if abs(len(word1) - len(word2)) > 2:
                    return float('inf')

                # Nếu giống nhau hoàn toàn
                if word1 == word2:
                    return 0

                # Tính số ký tự khác nhau (character-level comparison)
                # Cho phép swap, insert, delete (tối đa 2 operations)
                min_len = min(len(word1), len(word2))
                max_len = max(len(word1), len(word2))

                # Nếu độ dài khác nhau nhiều, không khớp
                if max_len - min_len > 2:
                    return float('inf')

                # Đếm số ký tự khác nhau ở các vị trí tương ứng
                diff = 0
                for i in range(min_len):
                    if word1[i] != word2[i]:
                        diff += 1

                # Thêm penalty cho độ dài khác nhau
                diff += (max_len - min_len)

                return diff

            # So sánh từng từ, cho phép 1-2 ký tự khác
            words_match = 0
            total_words = len(expected_words)

            if total_words > 0:
                for exp_word in expected_words:
                    # Tìm từ gần nhất trong actual_words
                    best_match = min(
                        [simple_edit_distance(exp_word, act_word) for act_word in actual_words],
                        default=float('inf')
                    )
                    # Nếu edit distance <= 2, coi như khớp
                    if best_match <= 2:
                        words_match += 1

                # Nếu >= 80% từ khớp, chấp nhận
                if words_match / total_words >= 0.8:
                    continue

            # Không khớp với bất kỳ cách nào
            missing_params.append(f'Cột {chr(65 + idx)}: Mong đợi "{expected_param}" nhưng có "{actual_value}"')

        if missing_params:
            return JsonResponse({
                'error': f'File Excel không hợp lệ: Header row 2 (tên thông số) không khớp với template Thông số H2. Các cột không hợp lệ:\n' + '\n'.join(missing_params[:5]) + (f'\n... và {len(missing_params) - 5} cột khác' if len(missing_params) > 5 else ''),
                'status': 'error',
                'missing_params': missing_params[:10]  # Chỉ trả về 10 đầu tiên
            }, status=400)

        # 4. Kiểm tra số hàng dữ liệu (từ row 4 đến row 27 phải có đủ 24 hàng)
        data_row_count = len(df) - 3  # Trừ 3 header rows
        if data_row_count < 24:
            return JsonResponse({
                'error': f'File Excel không hợp lệ: File thông số H2 phải có đúng 24 hàng dữ liệu (từ A4 đến A27). File hiện tại có {data_row_count} hàng dữ liệu.',
                'status': 'error',
                'expected_data_rows': 24,
                'actual_data_rows': data_row_count
            }, status=400)

        # 5. Cảnh báo nếu có quá nhiều hàng (có thể là file 48 slots)
        if data_row_count >= 48:
            return JsonResponse({
                'error': 'File Excel không hợp lệ: File có quá nhiều hàng dữ liệu (>48). Có thể đây là file Thông số Vận hành (48 slots) thay vì file Thông số H2 (24 slots). Vui lòng sử dụng file template Thông số H2.',
                'status': 'error',
                'detected_rows': data_row_count,
                'expected_rows': 24,
                'suggestion': 'Có thể bạn đang import nhầm file Thông số Vận hành'
            }, status=400)

        # ===== VALIDATION HOÀN TẤT =====

        # Lấy thiết bị H2
        device_code = request_data.get('device_code', 'SH.TB.H2.GE')  # Default to H2
        try:
            thiet_bi = filter_queryset_by_factory(
                ThietBi.objects.all(),
                request.user,
                'nha_may',
                'string',
            ).get(ma_day_du=device_code)
        except ThietBi.DoesNotExist:
            return HttpResponse(f'Không tìm thấy thiết bị {device_code}', status=400)

        # Định nghĩa mapping cho các cột (A đến T)
        column_mapping = {
            0: {"ten": "Áp lực nước", "ma": "ap_luc_nuoc", "don_vi": "bar"},
            1: {"ten": "Áp lực chèn trục", "ma": "ap_luc_chen_truc", "don_vi": "bar"},
            2: {"ten": "Lưu lượng chèn trục", "ma": "luu_luong_chen_truc", "don_vi": "l/p"},
            3: {"ten": "Lưu lượng ổ hướng tuabin", "ma": "luu_luong_o_huong_tuabin", "don_vi": "l/p"},
            4: {"ten": "Nhiệt độ ổ hướng tuabin", "ma": "nhiet_do_o_huong_tuabin", "don_vi": "°C"},
            5: {"ten": "Lưu lượng ổ hướng máy phát", "ma": "luu_luong_o_huong_may_phat", "don_vi": "l/p"},
            6: {"ten": "Nhiệt độ ổ hướng máy phát", "ma": "nhiet_do_o_huong_may_phat", "don_vi": "°C"},
            7: {"ten": "Lưu lượng ổ đỡ máy phát", "ma": "luu_luong_o_do_may_phat", "don_vi": "l/p"},
            8: {"ten": "Nhiệt độ ổ đỡ", "ma": "nhiet_do_o_do", "don_vi": "°C"},
            9: {"ten": "Nhiệt độ ổ hướng - ổ đỡ", "ma": "nhiet_do_o_huong_o_do", "don_vi": "°C"},
            10: {"ten": "Nhiệt độ đầu ổ đỡ", "ma": "nhiet_do_dau_o_do", "don_vi": "°C"},
            11: {"ten": "Lưu lượng làm mát máy phát", "ma": "luu_luong_lam_mat_may_phat", "don_vi": "l/p"},
            12: {"ten": "Nhiệt độ nước làm mát máy phát", "ma": "nhiet_do_nuoc_lam_mat_may_phat", "don_vi": "°C"},
            13: {"ten": "Nhiệt độ khí mát", "ma": "nhiet_do_khi_mat", "don_vi": "°C"},
            14: {"ten": "Nhiệt độ khí nóng", "ma": "nhiet_do_khi_nong", "don_vi": "°C"},
            15: {"ten": "Nhiệt độ cuộn dây stato", "ma": "nhiet_do_cuon_day_stato", "don_vi": "°C"},
            16: {"ten": "Tốc độ", "ma": "toc_do", "don_vi": "v/ph"},
            17: {"ten": "Giới hạn độ mở cánh hướng", "ma": "gioi_han_do_mo_canh_huong", "don_vi": "%"},
            18: {"ten": "Độ mở cánh hướng", "ma": "do_mo_canh_huong", "don_vi": "%"},
            19: {"ten": "Độ rơi tốc", "ma": "do_roi_toc", "don_vi": "%"},
        }

        # Tạo 24 mốc DateTime (00:00 -> 23:00) theo Asia/Ho_Chi_Minh
        import pytz
        vietnam_tz = pytz.timezone('Asia/Ho_Chi_Minh')
        time_cycles = []
        for hour in range(24):
            naive_dt = datetime.combine(target_date, time(hour, 0))
            time_cycles.append(vietnam_tz.localize(naive_dt))

        imported_count = 0

        # Lấy trước các bản ghi hiện có để tra cứu trên memory
        existing_records = ThongSoToMay.objects.filter(
            thiet_bi=thiet_bi,
            ngay_nhap=target_date
        )
        existing_lookup = {
            (rec.ten_thong_so, rec.thoi_diem_nhap): rec
            for rec in existing_records
        }

        to_create = []
        to_update = []

        # Xử lý dữ liệu từ A4 đến T27 (24 rows × 20 columns = 480 ô)
        for row_idx in range(3, 27):  # A4 to A27
            if row_idx >= len(df):
                break

            time_cycle = time_cycles[row_idx - 3] if row_idx - 3 < len(time_cycles) else None
            if not time_cycle:
                continue

            # Xử lý từng cột (A to T) - 20 cột
            for col_idx in range(0, 20):  # A to T (columns 0-19)
                if col_idx not in column_mapping:
                    continue

                mapping = column_mapping[col_idx]
                value = df.iloc[row_idx, col_idx]

                # Xử lý 3 dạng ô: có số, trống, dấu "-"
                if pd.isna(value) or value == '' or value == '-':
                    # Ô trống hoặc dấu "-" - import với giá trị null
                    numeric_value = None
                else:
                    try:
                        numeric_value = float(value)
                    except (ValueError, TypeError):
                        # Nếu không chuyển đổi được, dùng null
                        numeric_value = None

                # Tạo hoặc cập nhật thông số cho tất cả ô
                try:
                    nha_may_val = get_user_factory_name(request.user) or thiet_bi.nha_may
                    lookup_key = (mapping["ten"], time_cycle)

                    if lookup_key in existing_lookup:
                        obj = existing_lookup[lookup_key]
                        if obj.gia_tri != numeric_value or obj.nha_may != nha_may_val:
                            obj.gia_tri = numeric_value
                            obj.nha_may = nha_may_val
                            to_update.append(obj)
                    else:
                        obj = ThongSoToMay(
                            thiet_bi=thiet_bi,
                            ten_thong_so=mapping["ten"],
                            thoi_diem_nhap=time_cycle,
                            ngay_nhap=target_date,
                            ma_thong_so=mapping["ma"],
                            don_vi=mapping["don_vi"],
                            gia_tri=numeric_value,
                            nha_may=nha_may_val,
                            ky_hieu_van_hanh=f'{device_code.split(".")[-2]}_{mapping["ma"]}',
                            ghi_chu=f'Import từ Excel - {target_date}'
                        )
                        to_create.append(obj)
                        existing_lookup[lookup_key] = obj

                    imported_count += 1
                except Exception:
                    continue

        # Ghi hàng loạt xuống DB
        if to_create:
            ThongSoToMay.objects.bulk_create(to_create)
        if to_update:
            ThongSoToMay.objects.bulk_update(to_update, ['gia_tri', 'nha_may'])

        return JsonResponse({
            'message': f'Import thành công {imported_count} bản ghi',
            'imported_count': imported_count,
            'status': 'success'
        }, status=200)

    except Exception as e:
        return JsonResponse({
            'error': f'Lỗi khi import: {str(e)}',
            'status': 'error'
        }, status=500)

